#!/usr/bin/env python3
"""2022A 问题二：PTO 阻尼参数与平均输出功率优化主求解。

本脚本必须由用户本地 full_fidelity 执行。它只读取项目根目录中的附件3、附件4，
不读取网上答案，也不沿用问题一的水动力数值。输出：
    问题二求解/问题二求解结果.xlsx

模型语义绑定：
Q2 semantic revision=1
hash=513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862
"""
from __future__ import annotations

import hashlib
import math
import platform
from copy import copy
from pathlib import Path

import numpy as np
import scipy
from openpyxl import Workbook, load_workbook
from scipy.integrate import solve_ivp
from scipy.optimize import differential_evolution, minimize


SEMANTIC_REVISION = 1
SEMANTIC_HASH = "513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862"
EXPECTED_HASHES = {
    "附件3.xlsx": "50a5dd70f04dfb0a57fb2602422dc7999b30aad54ddc02353f5b8f01423fd612",
    "附件4.xlsx": "c8eff812f5980d955b4f0e587c5f7a357b2571d8d903fcb4913fba77c7354d6d",
}
LOCKED_DATA_SHA256 = "0cc51ac30576d2c2d3901aaf27dfe26526bd3768035c2685e9a57d4f81551653"
C_MAX = 100000.0
A_MAX = 100000.0
RTOL = 1.0e-10
ATOL = 1.0e-12
DE_SEED = 2022
DE_MAXITER = 30
DE_POPSIZE = 12
DE_TOL = 1.0e-5
SEARCH_TRANSIENT_CYCLES = 50
SEARCH_AVERAGE_CYCLES = 10
FINAL_PRE_CYCLES = 80
FINAL_BLOCK_CYCLES = 10
FINAL_MAX_CYCLES = 160
POWER_STABILITY_TOL = 1.0e-5
PHASE_STABILITY_TOL = 1.0e-5
LINEAR_CROSS_TOL = 1.0e-6
N0_EQUIV_TOL = 1.0e-8
STEP_REFINEMENT_TOL = 1.0e-6
NEIGHBOR_TOL = 2.0e-4

FULL_FIDELITY_CONFIG = {
    "execution_owner": "user",
    "execution_profile": "full_fidelity",
    "stage": "primary",
    "problem_name": "问题二",
    "semantic_revision": SEMANTIC_REVISION,
    "semantic_hash": SEMANTIC_HASH,
    "data_paths": ["附件3.xlsx", "附件4.xlsx"],
    "data_sha256": LOCKED_DATA_SHA256,
    "solver": "DOP853 + analytic linear optimum + differential_evolution + bounded Powell refinement",
    "solver_version": "SciPy runtime version is recorded in workbook",
    "random_seed": DE_SEED,
    "tolerance": RTOL,
    "iteration_or_time_limit": (
        "DE maxiter=30,popsize=12; each nonlinear search candidate uses 60T "
        "(last 10T power); final verification >=100T and extends to 160T if needed"
    ),
    "expected_workbook": "问题二求解/问题二求解结果.xlsx",
    "allow_reduced_data": False,
    "allow_coarser_grid": False,
    "allow_shorter_horizon": False,
    "allow_fewer_repetitions": False,
    "allow_relaxed_tolerance": False,
    "allow_silent_solver_fallback": False,
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def aggregate_data_hash(actual_hashes: dict[str, str]) -> str:
    text = "".join(f"{name}:{actual_hashes[name]}\n" for name in sorted(actual_hashes))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def audit_source_files(project_root: Path) -> tuple[dict[str, str], list[list[object]]]:
    actual: dict[str, str] = {}
    rows: list[list[object]] = []
    for name, expected in EXPECTED_HASHES.items():
        path = project_root / name
        if not path.is_file():
            raise FileNotFoundError(
                f"缺少输入文件: {path}。请把附件3.xlsx、附件4.xlsx直接放在A题根目录，"
                "与‘问题二求解’文件夹同级。"
            )
        current = sha256_file(path)
        actual[name] = current
        passed = current.lower() == expected.lower()
        rows.append(["通过" if passed else "阻断", f"{name} SHA-256", current, expected])
        if not passed:
            raise ValueError(f"{name} 哈希与2022A锁定附件不一致，停止求解")
    if aggregate_data_hash(actual) != LOCKED_DATA_SHA256:
        raise ValueError("Q2聚合data_sha256与锁定配置不一致")
    return actual, rows


def read_q2_parameters(project_root: Path) -> tuple[dict[str, float], list[list[object]]]:
    wb3 = load_workbook(project_root / "附件3.xlsx", data_only=True, read_only=True)
    ws3 = wb3.active
    expected3 = [
        "问题", "入射波浪频率 (s-1)", "垂荡附加质量 (kg)", "纵摇附加转动惯量 (kg·m2)",
        "垂荡兴波阻尼系数 (N·s/m)", "纵摇兴波阻尼系数 (N·m·s)",
        "垂荡激励力振幅 (N)", "纵摇激励力矩振幅 (N·m)",
    ]
    headers3 = [ws3.cell(1, col).value for col in range(1, 9)]
    if headers3 != expected3:
        raise ValueError("附件3表头与2022A锁定版本不一致")
    q2_row = next((row for row in ws3.iter_rows(min_row=2, values_only=True) if row[0] == "问题2"), None)
    if q2_row is None:
        raise ValueError("附件3中未找到问题2参数行")

    wb4 = load_workbook(project_root / "附件4.xlsx", data_only=True, read_only=True)
    ws4 = wb4.active
    if [ws4["A1"].value, ws4["B1"].value] != ["参数", "取值"]:
        raise ValueError("附件4表头与2022A锁定版本不一致")
    p4 = {str(row[0]): float(row[1]) for row in ws4.iter_rows(min_row=2, values_only=True) if row[0] is not None}
    needed = ["浮子质量 (kg)", "浮子底半径 (m)", "振子质量 (kg)", "海水的密度 (kg/m3)",
              "重力加速度 (m/s2)", "弹簧刚度 (N/m)"]
    if any(name not in p4 for name in needed):
        raise ValueError("附件4缺少Q2所需字段")

    params = {
        "omega": float(q2_row[1]),
        "added_mass": float(q2_row[2]),
        "wave_damping": float(q2_row[4]),
        "force_amp": float(q2_row[6]),
        "float_mass": p4["浮子质量 (kg)"],
        "radius": p4["浮子底半径 (m)"],
        "osc_mass": p4["振子质量 (kg)"],
        "rho": p4["海水的密度 (kg/m3)"],
        "g": p4["重力加速度 (m/s2)"],
        "spring_k": p4["弹簧刚度 (N/m)"],
    }
    if not all(math.isfinite(value) for value in params.values()):
        raise ValueError("Q2参数存在NaN或Inf")
    if not math.isclose(params["omega"], 2.2143, rel_tol=0.0, abs_tol=1.0e-12):
        raise ValueError("附件3问题2入射波浪频率不是锁定值2.2143 s^-1")
    rows = [
        ["通过", "附件3问题2参数", "ω、垂荡附加质量、垂荡兴波阻尼、激励力均直接读取问题2行", "不沿用Q1数值"],
        ["通过", "附件4字段", "Q2所需质量/半径/密度/g/弹簧刚度齐全", "原始值直接使用"],
        ["通过", "预处理判定", "not_needed", "仅非破坏性字段/单位/NaN/Inf审计"],
    ]
    return params, rows


def hydro_stiffness(params: dict[str, float]) -> float:
    return params["rho"] * params["g"] * math.pi * params["radius"] ** 2


def damper_force(vr: float, damping: tuple[str, float, float]) -> float:
    mode, p1, p2 = damping
    if mode == "linear":
        return p1 * vr
    if mode == "power":
        return p1 * abs(vr) ** p2 * vr
    raise ValueError(f"未知阻尼模式: {mode}")


def instant_power(vr: float, damping: tuple[str, float, float]) -> float:
    mode, p1, p2 = damping
    if mode == "linear":
        return p1 * vr * vr
    if mode == "power":
        return p1 * abs(vr) ** (p2 + 2.0)
    raise ValueError(f"未知阻尼模式: {mode}")


def state_rhs(params: dict[str, float], damping: tuple[str, float, float]):
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    kh = hydro_stiffness(params)
    b = params["wave_damping"]
    force_amp = params["force_amp"]
    omega = params["omega"]

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        zf, vf, zo, vo = y
        vr = vf - vo
        fd = damper_force(vr, damping)
        af = (force_amp * math.cos(omega * t) - b * vf - kh * zf - k * (zf - zo) - fd) / mass
        ao = (k * (zf - zo) + fd) / mo
        return np.array([vf, af, vo, ao], dtype=float)

    return rhs


def integrate_state(
    params: dict[str, float], damping: tuple[str, float, float],
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> np.ndarray:
    sol = solve_ivp(
        state_rhs(params, damping), (t0, t1), y0, method="DOP853",
        rtol=RTOL, atol=ATOL, max_step=max_step,
    )
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"状态积分失败: {sol.message}")
    return np.asarray(sol.y[:, -1], dtype=float)


def integrate_power_window(
    params: dict[str, float], damping: tuple[str, float, float],
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> tuple[np.ndarray, float]:
    base_rhs = state_rhs(params, damping)

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        dy = base_rhs(t, y[:4])
        vr = float(y[1] - y[3])
        return np.array([*dy, instant_power(vr, damping)], dtype=float)

    initial = np.array([*y0, 0.0], dtype=float)
    sol = solve_ivp(rhs, (t0, t1), initial, method="DOP853", rtol=RTOL, atol=ATOL, max_step=max_step)
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"功率积分失败: {sol.message}")
    end = np.asarray(sol.y[:4, -1], dtype=float)
    average_power = float(sol.y[4, -1] / (t1 - t0))
    if average_power < -1.0e-10 or not math.isfinite(average_power):
        raise RuntimeError("平均PTO功率出现非有限值或明显负值")
    return end, max(0.0, average_power)


def linear_power_coefficients(params: dict[str, float]) -> dict[str, float]:
    omega = params["omega"]
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    h = hydro_stiffness(params) - mass * omega ** 2
    bomega = params["wave_damping"] * omega
    gterm = mo * omega ** 2 - k
    a0 = h * gterm + mo * omega ** 2 * k
    b0 = bomega * gterm
    a1 = bomega * omega
    b1 = omega * (mo * omega ** 2 - h)
    alpha = a1 * a1 + b1 * b1
    beta = 2.0 * (a0 * a1 + b0 * b1)
    gamma = a0 * a0 + b0 * b0
    numerator_factor = 0.5 * omega ** 2 * params["force_amp"] ** 2 * (mo * omega ** 2) ** 2
    if alpha <= 0.0 or gamma <= 0.0:
        raise RuntimeError("直线阻尼解析功率系数退化，无法使用当前闭式最优公式")
    return {
        "H": h, "B": bomega, "G": gterm, "A0": a0, "B0": b0, "A1": a1, "B1": b1,
        "alpha": alpha, "beta": beta, "gamma": gamma, "numerator_factor": numerator_factor,
    }


def linear_power_closed_form(c_value: float, coeff: dict[str, float]) -> float:
    denominator = coeff["alpha"] * c_value ** 2 + coeff["beta"] * c_value + coeff["gamma"]
    if denominator <= 0.0:
        raise RuntimeError("直线阻尼功率解析式分母非正")
    return coeff["numerator_factor"] * c_value / denominator


def linear_optimum(params: dict[str, float]) -> tuple[float, float, dict[str, float]]:
    coeff = linear_power_coefficients(params)
    c0 = math.sqrt(coeff["gamma"] / coeff["alpha"])
    c_star = min(C_MAX, max(0.0, c0))
    p_star = linear_power_closed_form(c_star, coeff)
    endpoint_best = max(linear_power_closed_form(0.0, coeff), linear_power_closed_form(C_MAX, coeff))
    if p_star + 1.0e-10 < endpoint_best:
        raise RuntimeError("解析驻点未优于区间端点，直线阻尼最优性检查失败")
    return c_star, p_star, coeff


def search_power(params: dict[str, float], a_value: float, n_value: float) -> float:
    period = 2.0 * math.pi / params["omega"]
    damping = ("power", float(a_value), float(n_value))
    y = integrate_state(params, damping, 0.0, SEARCH_TRANSIENT_CYCLES * period, np.zeros(4), period / 20.0)
    _, power = integrate_power_window(
        params, damping, SEARCH_TRANSIENT_CYCLES * period,
        (SEARCH_TRANSIENT_CYCLES + SEARCH_AVERAGE_CYCLES) * period, y, period / 20.0,
    )
    return power


def verify_periodic_candidate(
    params: dict[str, float], damping: tuple[str, float, float], step_divisor: float,
) -> dict[str, object]:
    period = 2.0 * math.pi / params["omega"]
    max_step = period / step_divisor
    y = integrate_state(params, damping, 0.0, FINAL_PRE_CYCLES * period, np.zeros(4), max_step)
    previous_end: np.ndarray | None = None
    previous_power: float | None = None
    diagnostics: list[list[object]] = []
    cycle = FINAL_PRE_CYCLES
    while cycle < FINAL_MAX_CYCLES:
        next_cycle = cycle + FINAL_BLOCK_CYCLES
        end, power = integrate_power_window(params, damping, cycle * period, next_cycle * period, y, max_step)
        if previous_end is not None and previous_power is not None:
            power_rel = abs(power - previous_power) / (1.0 + abs(power))
            phase_rel = float(np.linalg.norm(end - previous_end) / (1.0 + np.linalg.norm(end)))
            passed = power_rel <= POWER_STABILITY_TOL and phase_rel <= PHASE_STABILITY_TOL
            diagnostics.append([cycle - FINAL_BLOCK_CYCLES, cycle, next_cycle, previous_power, power, power_rel, phase_rel, passed])
            if passed:
                return {"power": power, "end_state": end, "stop_cycle": next_cycle, "diagnostics": diagnostics}
        previous_end = end.copy()
        previous_power = power
        y = end
        cycle = next_cycle
    raise RuntimeError("候选点在160T内未满足功率窗口与同相位状态稳定性阈值")


def nonlinear_optimum(params: dict[str, float], c_linear: float) -> tuple[dict[str, object], list[list[object]]]:
    objective_calls = 0

    def objective(x: np.ndarray) -> float:
        nonlocal objective_calls
        objective_calls += 1
        a_value = A_MAX * float(x[0])
        n_value = float(x[1])
        return -search_power(params, a_value, n_value)

    x0 = np.array([c_linear / A_MAX, 0.0], dtype=float)
    de = differential_evolution(
        objective, bounds=[(0.0, 1.0), (0.0, 1.0)], seed=DE_SEED,
        maxiter=DE_MAXITER, popsize=DE_POPSIZE, tol=DE_TOL, atol=0.0,
        polish=False, updating="immediate", workers=1, x0=x0,
    )
    local = minimize(
        objective, np.clip(de.x, 0.0, 1.0), method="Powell",
        bounds=[(0.0, 1.0), (0.0, 1.0)],
        options={"xtol": 1.0e-6, "ftol": 1.0e-8, "maxiter": 30, "disp": False},
    )
    if not de.success and not local.success:
        raise RuntimeError(f"DE与局部精修均未正常终止: DE={de.message}; Powell={local.message}")
    candidates = [(np.asarray(de.x, dtype=float), float(de.fun), "DE")]
    if np.all(np.isfinite(local.x)) and math.isfinite(float(local.fun)):
        candidates.append((np.asarray(local.x, dtype=float), float(local.fun), "DE+Powell"))
    best_x, best_fun, source = min(candidates, key=lambda item: item[1])
    best_x = np.clip(best_x, 0.0, 1.0)
    result = {
        "u": float(best_x[0]), "a": A_MAX * float(best_x[0]), "n": float(best_x[1]),
        "search_power": -best_fun, "source": source, "de_success": bool(de.success),
        "de_message": str(de.message), "de_nit": int(de.nit), "de_nfev": int(de.nfev),
        "local_success": bool(local.success), "local_message": str(local.message),
        "local_nit": int(getattr(local, "nit", -1)), "local_nfev": int(getattr(local, "nfev", -1)),
        "objective_calls": objective_calls,
    }
    rows = [
        ["differential_evolution", de.success, de.message, de.nit, de.nfev, float(de.x[0]), float(de.x[1]), -float(de.fun)],
        ["bounded Powell refinement", local.success, local.message, getattr(local, "nit", None),
         getattr(local, "nfev", None), float(local.x[0]), float(local.x[1]), -float(local.fun)],
        ["selected", True, source, None, objective_calls, result["u"], result["n"], result["search_power"]],
    ]
    return result, rows


def neighborhood_check(
    params: dict[str, float], result: dict[str, object], c_linear: float,
) -> tuple[list[list[object]], float]:
    u = float(result["u"])
    n = float(result["n"])
    candidate_power = search_power(params, A_MAX * u, n)
    raw_points = [
        ("u-0.01", u - 0.01, n), ("u+0.01", u + 0.01, n),
        ("n-0.01", u, n - 0.01), ("n+0.01", u, n + 0.01),
        ("diag--", u - 0.01, n - 0.01), ("diag++", u + 0.01, n + 0.01),
        ("a=0", 0.0, n), ("a=100000", 1.0, n),
        ("n=0 linear-boundary", c_linear / A_MAX, 0.0), ("n=1 projection", u, 1.0),
    ]
    rows: list[list[object]] = []
    best_alt = -math.inf
    seen: set[tuple[float, float]] = set()
    for label, uu, nn in raw_points:
        uu = min(1.0, max(0.0, float(uu)))
        nn = min(1.0, max(0.0, float(nn)))
        key = (round(uu, 12), round(nn, 12))
        if key in seen or (math.isclose(uu, u, abs_tol=1.0e-12) and math.isclose(nn, n, abs_tol=1.0e-12)):
            continue
        seen.add(key)
        power = search_power(params, A_MAX * uu, nn)
        best_alt = max(best_alt, power)
        rows.append([label, A_MAX * uu, nn, power, power - candidate_power])
    if best_alt > candidate_power * (1.0 + NEIGHBOR_TOL) + 1.0e-8:
        raise RuntimeError("边界/局部邻域发现显著高于当前非线性候选的功率，优化结果不能验收")
    return rows, candidate_power


def add_sheet(book: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(46, max(11, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def write_workbook(
    output: Path, code_hash: str, params: dict[str, float], audit_rows: list[list[object]],
    linear: dict[str, object], nonlinear: dict[str, object], opt_rows: list[list[object]],
    convergence_rows: list[list[object]], neighbor_rows: list[list[object]], quality_rows: list[list[object]],
) -> None:
    book = Workbook()
    book.remove(book.active)
    runtime_rows = [[key, value] for key, value in FULL_FIDELITY_CONFIG.items()]
    runtime_rows += [["code_sha256", code_hash], ["scipy_version", scipy.__version__], ["platform", platform.platform()],
                     ["fallback_used", False], ["actual_stop_reason", "all Q2 quality gates passed"]]
    add_sheet(book, "运行配置", ["项目", "值"], runtime_rows)
    add_sheet(book, "数据审计", ["状态", "检查项", "实际/结论", "期望/处理"], audit_rows)

    core_rows = [
        ["问题2入射波浪频率ω", params["omega"], "s^-1"],
        ["直线阻尼最优c", linear["c_star"], "N·s/m"],
        ["直线阻尼最大平均功率", linear["closed_power"], "W"],
        ["直线阻尼时域核验平均功率", linear["time_power"], "W"],
        ["幂律阻尼最优比例系数a", nonlinear["a"], "N·(s/m)^(n+1)"],
        ["幂律阻尼最优幂指数n", nonlinear["n"], "1"],
        ["幂律阻尼最大平均功率", nonlinear["final_power"], "W"],
        ["幂律候选最终验证停止周期", nonlinear["stop_cycle"], "T"],
    ]
    add_sheet(book, "核心指标", ["指标", "数值", "单位"], core_rows)

    coeff = linear["coeff"]
    linear_rows = [
        ["H", coeff["H"], ""], ["B", coeff["B"], ""], ["G", coeff["G"], ""],
        ["A0", coeff["A0"], ""], ["B0", coeff["B0"], ""], ["A1", coeff["A1"], ""], ["B1", coeff["B1"], ""],
        ["alpha", coeff["alpha"], ""], ["beta", coeff["beta"], ""], ["gamma", coeff["gamma"], ""],
        ["c0=sqrt(gamma/alpha)", linear["c0"], "N·s/m"], ["c*=clip(c0)", linear["c_star"], "N·s/m"],
        ["Pbar_closed(c*)", linear["closed_power"], "W"], ["Pbar_time(c*)", linear["time_power"], "W"],
        ["解析-时域相对差", linear["cross_error"], "1"],
    ]
    add_sheet(book, "直线阻尼解析", ["项目", "数值", "单位"], linear_rows)
    add_sheet(book, "优化诊断", ["阶段", "success", "message/source", "nit", "nfev/calls", "u=a/100000", "n", "search_power/W"], opt_rows)
    add_sheet(book, "收敛诊断", ["对象", "前窗口起始周期", "前窗口结束周期", "后窗口结束周期",
                                  "前窗口功率/W", "后窗口功率/W", "功率相对差", "同相位状态相对差", "通过"], convergence_rows)
    add_sheet(book, "边界邻域检查", ["检查点", "a", "n", "60T搜索口径功率/W", "相对候选差/W"], neighbor_rows)
    add_sheet(book, "主结果质量门", ["检查项", "是否通过", "证据"], quality_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)


def main() -> None:
    script_path = Path(__file__).resolve()
    project_root = script_path.parents[1]
    output = project_root / "问题二求解" / "问题二求解结果.xlsx"
    _, audit_rows = audit_source_files(project_root)
    params, parameter_rows = read_q2_parameters(project_root)
    audit_rows.extend(parameter_rows)
    c_star, closed_power, coeff = linear_optimum(params)
    c0 = math.sqrt(coeff["gamma"] / coeff["alpha"])
    linear_verify = verify_periodic_candidate(params, ("linear", c_star, 0.0), 40.0)
    linear_time_power = float(linear_verify["power"])
    linear_cross_error = abs(linear_time_power - closed_power) / (1.0 + abs(closed_power))
    if linear_cross_error > LINEAR_CROSS_TOL:
        raise RuntimeError("直线阻尼解析功率与DOP853稳态功率不一致")

    nonlinear, opt_rows = nonlinear_optimum(params, c_star)
    nonlinear_verify = verify_periodic_candidate(params, ("power", float(nonlinear["a"]), float(nonlinear["n"])), 40.0)
    refined_verify = verify_periodic_candidate(params, ("power", float(nonlinear["a"]), float(nonlinear["n"])), 80.0)
    step_error = abs(float(refined_verify["power"]) - float(nonlinear_verify["power"])) / (1.0 + abs(float(refined_verify["power"])))
    if step_error > STEP_REFINEMENT_TOL:
        raise RuntimeError("非线性最优候选步长加密复算未通过")

    n0_verify = verify_periodic_candidate(params, ("power", c_star, 0.0), 40.0)
    n0_error = abs(float(n0_verify["power"]) - linear_time_power) / (1.0 + abs(linear_time_power))
    if n0_error > N0_EQUIV_TOL:
        raise RuntimeError("n=0时幂律模型未严格退化为直线阻尼模型")
    if float(nonlinear_verify["power"]) + 1.0e-6 < linear_time_power * (1.0 - NEIGHBOR_TOL):
        raise RuntimeError("非线性二维优化结果低于其合法n=0边界最优值，停止验收")

    neighbor_rows, candidate_search_power = neighborhood_check(params, nonlinear, c_star)
    nonlinear["final_power"] = float(nonlinear_verify["power"])
    nonlinear["refined_power"] = float(refined_verify["power"])
    nonlinear["step_error"] = step_error
    nonlinear["stop_cycle"] = int(nonlinear_verify["stop_cycle"])
    nonlinear["candidate_search_power"] = candidate_search_power

    convergence_rows: list[list[object]] = []
    for label, result in [
        ("直线c*", linear_verify), ("幂律候选", nonlinear_verify),
        ("幂律候选-步长加密", refined_verify), ("n=0等价", n0_verify),
    ]:
        for row in result["diagnostics"]:
            convergence_rows.append([label, *row])

    quality_rows = [
        ["原始附件哈希与Q2数据作用域", "是", "附件3/4 SHA-256与锁定版本一致，且读取附件3问题2行"],
        ["直线阻尼区间最优性", "是", f"c0={c0:.10f}, c*={c_star:.10f}; 已比较区间端点"],
        ["直线解析-时域交叉核验", "是", f"相对差={linear_cross_error:.3e} <= {LINEAR_CROSS_TOL:.1e}"],
        ["幂律候选参数范围", "是", f"a={nonlinear['a']:.10f}, n={nonlinear['n']:.10f} 均在题面范围内"],
        ["幂律周期稳态收敛", "是", f"最终验证至{nonlinear['stop_cycle']}T满足功率/同相位状态阈值"],
        ["幂律步长加密", "是", f"相对功率差={step_error:.3e} <= {STEP_REFINEMENT_TOL:.1e}"],
        ["n=0跨模型等价", "是", f"相对差={n0_error:.3e} <= {N0_EQUIV_TOL:.1e}"],
        ["n=0下界检查", "是", "二维幂律候选功率不低于合法线性边界最优值（数值容差内）"],
        ["边界/局部邻域检查", "是", f"{len(neighbor_rows)}个邻域/边界候选未发现超过当前候选{NEIGHBOR_TOL:.2%}的点"],
    ]

    linear = {
        "c0": c0, "c_star": c_star, "closed_power": closed_power, "time_power": linear_time_power,
        "cross_error": linear_cross_error, "coeff": coeff,
    }
    code_hash = sha256_file(script_path)
    write_workbook(output, code_hash, params, audit_rows, linear, nonlinear, opt_rows, convergence_rows, neighbor_rows, quality_rows)


if __name__ == "__main__":
    main()
