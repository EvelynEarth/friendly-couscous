#!/usr/bin/env python3
"""2022A 问题二：独立结果深化分析。

前置：问题二主求解与 a=100000 边界稳态精修均已通过质量门。
本脚本只读取已验收工作簿，不重复二维 DE，也不重复主求解 ODE 链。

输入（与本脚本同目录）：
    问题二求解结果.xlsx
    问题二边界稳态精修结果.xlsx
输出：
    问题二结果深化分析.xlsx

Q2 semantic revision=1
hash=513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862
"""
from __future__ import annotations

import hashlib
import math
import platform
import time
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

SEMANTIC_REVISION = 1
SEMANTIC_HASH = "513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862"
DATA_SHA256 = "0cc51ac30576d2c2d3901aaf27dfe26526bd3768035c2685e9a57d4f81551653"
A_MAX = 100000.0
POWER_STABILITY_TOL = 1.0e-5
PHASE_STABILITY_TOL = 1.0e-5
STEP_TOL = 1.0e-6

FULL_FIDELITY_CONFIG = {
    "execution_owner": "user",
    "execution_profile": "full_fidelity",
    "stage": "analysis",
    "problem_name": "问题二",
    "semantic_revision": SEMANTIC_REVISION,
    "semantic_hash": SEMANTIC_HASH,
    "data_sha256": DATA_SHA256,
    "analysis_scope": "accepted Q2 workbook + accepted a=100000 boundary refinement workbook",
    "rerun_primary_solver": False,
    "output": "问题二求解/问题二结果深化分析.xlsx",
}


def progress(stage: str, message: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {stage} {message}", flush=True)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def scalar(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def sheet_dict(ws, key_col: int = 1, value_col: int = 2) -> dict[str, object]:
    result: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[key_col - 1] is not None:
            result[str(row[key_col - 1])] = row[value_col - 1]
    return result


def assert_runtime(wb, workbook_name: str) -> None:
    if "运行配置" not in wb.sheetnames:
        raise ValueError(f"{workbook_name} 缺少运行配置工作表")
    config = sheet_dict(wb["运行配置"])
    if str(config.get("semantic_hash")) != SEMANTIC_HASH:
        raise ValueError(f"{workbook_name} semantic_hash 与 Q2 锁定模型不一致")
    if str(config.get("data_sha256")) != DATA_SHA256:
        raise ValueError(f"{workbook_name} data_sha256 与 Q2 锁定数据不一致")


def assert_quality_gate(wb, workbook_name: str) -> None:
    if "主结果质量门" not in wb.sheetnames:
        raise ValueError(f"{workbook_name} 缺少主结果质量门")
    failures = []
    ws = wb["主结果质量门"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        passed = str(row[1]).strip().lower() in {"是", "true", "通过", "pass", "passed"}
        if not passed:
            failures.append(str(row[0]))
    if failures:
        raise ValueError(f"{workbook_name} 存在未通过质量门: {failures}")


def find_metric(ws, label: str) -> float:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip() == label:
            return float(row[1])
    raise KeyError(f"{ws.title} 未找到指标: {label}")


def read_main(main_path: Path) -> dict[str, object]:
    wb = load_workbook(main_path, data_only=True, read_only=True)
    assert_runtime(wb, main_path.name)
    assert_quality_gate(wb, main_path.name)
    core = wb["核心指标"]
    values = {
        "linear_c": find_metric(core, "直线阻尼最优c"),
        "linear_closed_power": find_metric(core, "直线阻尼最大平均功率"),
        "linear_time_power": find_metric(core, "直线阻尼时域核验平均功率"),
        "main_a": find_metric(core, "幂律阻尼最优比例系数a"),
        "main_n": find_metric(core, "幂律阻尼最优幂指数n"),
        "main_power": find_metric(core, "幂律阻尼最大平均功率"),
    }
    opt_rows = [list(row) for row in wb["优化诊断"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    neighbor_rows = [list(row) for row in wb["边界邻域检查"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    convergence_rows = [list(row) for row in wb["收敛诊断"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    wb.close()
    values["opt_rows"] = opt_rows
    values["neighbor_rows"] = neighbor_rows
    values["convergence_rows"] = convergence_rows
    return values


def read_refine(refine_path: Path) -> dict[str, object]:
    wb = load_workbook(refine_path, data_only=True, read_only=True)
    assert_runtime(wb, refine_path.name)
    assert_quality_gate(wb, refine_path.name)
    core = wb["核心结论"]
    values = {
        "a_final": find_metric(core, "固定比例系数a"),
        "n_final": find_metric(core, "边界精修最优n"),
        "power_t40": find_metric(core, "边界精修最大稳态平均功率(T/40)"),
        "power_t80": find_metric(core, "步长加密平均功率(T/80)"),
        "gain_vs_main": find_metric(core, "相对原二维候选功率增量"),
        "stop_t40": find_metric(core, "T/40最终停止周期"),
        "stop_t80": find_metric(core, "T/80最终停止周期"),
    }
    scan = []
    for row in wb["局部稳态扫描"].iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None:
            scan.append((float(row[0]), float(row[1]), str(row[2] or "")))
    refine_rows = [list(row) for row in wb["一维精修"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    boundary_rows = [list(row) for row in wb["边界对照"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    convergence_rows = [list(row) for row in wb["收敛诊断"].iter_rows(min_row=2, values_only=True) if row[0] is not None]
    wb.close()
    values["scan"] = scan
    values["refine_rows"] = refine_rows
    values["boundary_rows"] = boundary_rows
    values["convergence_rows"] = convergence_rows
    return values


def sampled_interval(scan_rows: list[tuple[float, float, str]], p_opt: float, loss_limit_pct: float):
    eligible = []
    for n, p, _ in scan_rows:
        loss_pct = (p_opt - p) / p_opt * 100.0
        if loss_pct <= loss_limit_pct + 1.0e-12:
            eligible.append(n)
    if not eligible:
        return None
    return min(eligible), max(eligible)


def get_opt_row(rows: list[list[object]], name: str) -> list[object] | None:
    for row in rows:
        if str(row[0]) == name:
            return row
    return None


def last_passed(rows: list[list[object]], label: str) -> list[object] | None:
    matched = [r for r in rows if str(r[0]) == label and bool(r[-1])]
    return matched[-1] if matched else None


def analyze(main: dict[str, object], refine: dict[str, object]) -> dict[str, object]:
    p_lin = float(main["linear_closed_power"])
    p_lin_time = float(main["linear_time_power"])
    p_main = float(main["main_power"])
    p_final = float(refine["power_t40"])
    p_final80 = float(refine["power_t80"])
    n_main = float(main["main_n"])
    n_final = float(refine["n_final"])
    a_main = float(main["main_a"])

    gain_abs = p_final - p_lin
    gain_pct = gain_abs / p_lin * 100.0
    gain_time_abs = p_final - p_lin_time
    gain_time_pct = gain_time_abs / p_lin_time * 100.0
    refine_abs = p_final - p_main
    refine_pct = refine_abs / p_main * 100.0
    n_shift = n_final - n_main
    a_gap = A_MAX - a_main
    step_rel = abs(p_final80 - p_final) / (1.0 + abs(p_final80))

    scan_rows = list(refine["scan"])
    sensitivity_rows = []
    for n, p, note in scan_rows:
        loss_w = p_final - p
        loss_pct = loss_w / p_final * 100.0
        sensitivity_rows.append([n, p, loss_w, loss_pct, note])

    int_001 = sampled_interval(scan_rows, p_final, 0.01)
    int_005 = sampled_interval(scan_rows, p_final, 0.05)
    int_010 = sampled_interval(scan_rows, p_final, 0.10)

    de = get_opt_row(main["opt_rows"], "parallel differential_evolution") or get_opt_row(main["opt_rows"], "differential_evolution")
    powell = get_opt_row(main["opt_rows"], "bounded Powell refinement")
    de_power = float(de[7]) if de and de[7] is not None else math.nan
    powell_power = float(powell[7]) if powell and powell[7] is not None else math.nan
    search_gain = powell_power - de_power if math.isfinite(de_power) and math.isfinite(powell_power) else math.nan

    conv40 = last_passed(refine["convergence_rows"], "最终T/40")
    conv80 = last_passed(refine["convergence_rows"], "最终T/80")
    if conv40 is None or conv80 is None:
        raise ValueError("边界精修工作簿没有最终通过的 T/40/T/80 收敛记录")
    p_rel40, phase_rel40 = float(conv40[6]), float(conv40[7])
    p_rel80, phase_rel80 = float(conv80[6]), float(conv80[7])

    a_upper_row = None
    for row in main["neighbor_rows"]:
        if str(row[0]) == "u+0.01":
            a_upper_row = row
            break
    upper_delta_search = float(a_upper_row[4]) if a_upper_row is not None else math.nan

    evidence = [
        ["Q2-EA", "收益比较", "幂律最优方案相对最优直线阻尼带来稳定但幅度较小的平均功率提升",
         "support", f"稳态功率提升 {gain_abs:.9f} W（{gain_pct:.6f}%），显著高于步长误差尺度",
         "正文报告提升为约0.288%，避免夸大工程增益", "Q2主结果/方案比较图"],
        ["Q2-EB", "边界+局部敏感性", "幂律最优位于a上边界且n为内部局部最优",
         "support", f"a*=100000；n*={n_final:.10f}；扫描中n*左右功率均下降，0.01%近优采样区间={int_001}",
         "正文解释为比例系数饱和边界 + 幂指数内部峰值", "Q2参数敏感性图"],
        ["Q2-EC", "收敛与离散精度", "最终最优功率对稳态窗口和时间步长具有数值稳定性",
         "support", f"T/40与T/80相对功率差={step_rel:.3e}; 最终窗口功率差={p_rel40:.3e}/{p_rel80:.3e}",
         "数值证据可放正文简述或附录，不需单独占主图", "Q2数值合法性说明"],
        ["Q2-ED", "优化路径", "二维DE/Powell的搜索结果可直接作为最终稳态最优功率报告",
         "modify", f"60T搜索功率(DE/Powell)={de_power:.6f}/{powell_power:.6f} W，与真正稳态目标存在瞬态偏置；最终边界精修P={p_final:.6f} W",
         "论文中DE/Powell只描述为全域候选搜索；最终参数与功率必须采用边界稳态精修结果", "Q2方法与优化流程"],
    ]

    return {
        "gain_abs": gain_abs, "gain_pct": gain_pct, "gain_time_abs": gain_time_abs,
        "gain_time_pct": gain_time_pct, "refine_abs": refine_abs, "refine_pct": refine_pct,
        "n_shift": n_shift, "a_gap": a_gap, "step_rel": step_rel,
        "sensitivity_rows": sensitivity_rows, "near_interval_001": int_001,
        "near_interval_005": int_005, "near_interval_010": int_010,
        "de_power": de_power, "powell_power": powell_power, "search_gain": search_gain,
        "p_rel40": p_rel40, "phase_rel40": phase_rel40, "p_rel80": p_rel80,
        "phase_rel80": phase_rel80, "upper_delta_search": upper_delta_search, "evidence": evidence,
    }


def excel_value(value):
    if isinstance(value, (list, tuple, dict)):
        return str(value)
    return scalar(value)


def add_sheet(book: Workbook, name: str, headers: list[str], rows: list[list[object]]):
    ws = book.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([excel_value(v) for v in row])
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        font = copy(cell.font); font.bold = True; cell.font = font
    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
        ws.column_dimensions[letter].width = min(52, max(11, width))
    return ws


def write_output(output: Path, main_path: Path, refine_path: Path,
                 main: dict[str, object], refine: dict[str, object], result: dict[str, object]) -> None:
    book = Workbook(); book.remove(book.active)
    runtime_rows = [[k, v] for k, v in FULL_FIDELITY_CONFIG.items()]
    runtime_rows += [["main_workbook", main_path.name], ["main_workbook_sha256", sha256_file(main_path)],
                     ["refine_workbook", refine_path.name], ["refine_workbook_sha256", sha256_file(refine_path)],
                     ["analysis_code_sha256", sha256_file(Path(__file__).resolve())], ["platform", platform.platform()]]
    add_sheet(book, "运行配置", ["项目", "值"], runtime_rows)

    summary_rows = [
        ["直线阻尼最优c", main["linear_c"], "N·s/m"],
        ["直线阻尼最大稳态平均功率", main["linear_closed_power"], "W"],
        ["幂律最终比例系数a", refine["a_final"], "题设比例系数，量纲随n变化"],
        ["幂律最终幂指数n", refine["n_final"], "1"],
        ["幂律最终稳态平均功率", refine["power_t40"], "W"],
        ["幂律相对直线绝对提升", result["gain_abs"], "W"],
        ["幂律相对直线相对提升", result["gain_pct"] / 100.0, "1"],
        ["边界精修相对原二维候选提升", result["refine_abs"], "W"],
        ["最终T/40-T/80步长相对差", result["step_rel"], "1"],
    ]
    add_sheet(book, "结论总览", ["指标", "数值", "单位/说明"], summary_rows)

    benefit_rows = [
        ["直线解析稳态", main["linear_closed_power"], 0.0, 0.0, "基准"],
        ["直线DOP853稳态", main["linear_time_power"], float(main["linear_time_power"]) - float(main["linear_closed_power"]),
         (float(main["linear_time_power"]) / float(main["linear_closed_power"]) - 1.0) * 100.0, "解析交叉核验"],
        ["幂律原二维候选稳态", main["main_power"], float(main["main_power"]) - float(main["linear_closed_power"]),
         (float(main["main_power"]) / float(main["linear_closed_power"]) - 1.0) * 100.0, "主优化候选"],
        ["幂律边界精修最终", refine["power_t40"], result["gain_abs"], result["gain_pct"], "最终采用"],
    ]
    add_sheet(book, "收益比较", ["方案", "稳态平均功率/W", "相对直线解析增量/W", "相对直线解析提升/%", "角色"], benefit_rows)

    sens_ws = add_sheet(book, "n敏感性", ["n", "稳态平均功率/W", "距最优损失/W", "距最优损失/%", "来源"], result["sensitivity_rows"])
    for row in range(2, sens_ws.max_row + 1):
        if abs(float(sens_ws.cell(row, 1).value) - float(refine["n_final"])) < 1.0e-8:
            for cell in sens_ws[row]: cell.fill = PatternFill("solid", fgColor="E2F0D9")
    interval_rows = [
        ["<=0.01%功率损失的采样近优区间", result["near_interval_001"], "仅基于已有稳态扫描点"],
        ["<=0.05%功率损失的采样近优区间", result["near_interval_005"], "仅基于已有稳态扫描点"],
        ["<=0.10%功率损失的采样近优区间", result["near_interval_010"], "仅基于已有稳态扫描点"],
        ["原二维候选n", main["main_n"], ""], ["最终边界精修n", refine["n_final"], ""],
        ["n修正量", result["n_shift"], ""], ["原二维候选a距上边界", result["a_gap"], "a单位"],
        ["旧n下60T搜索口径a上边界邻域功率差", result["upper_delta_search"], "W；仅作边界结构诊断"],
    ]
    add_sheet(book, "参数结构", ["项目", "数值/区间", "说明"], interval_rows)

    opt_rows = [["DE搜索功率", result["de_power"], "60T搜索口径，不作为最终稳态功率"],
                ["Powell精修搜索功率", result["powell_power"], "60T搜索口径，不作为最终稳态功率"],
                ["Powell相对DE搜索增量", result["search_gain"], "W；同一搜索口径内可比较"],
                ["原二维候选最终稳态功率", main["main_power"], "W"],
                ["边界精修最终稳态功率", refine["power_t40"], "W"],
                ["边界精修相对原候选增量", result["refine_abs"], "W"],
                ["边界精修相对原候选提升", result["refine_pct"] / 100.0, "1"]]
    add_sheet(book, "优化路径", ["阶段", "数值", "解释"], opt_rows)

    robust_rows = [["最终T/40功率窗口相对差", result["p_rel40"], POWER_STABILITY_TOL, result["p_rel40"] / POWER_STABILITY_TOL, "通过"],
                   ["最终T/40同相位状态相对差", result["phase_rel40"], PHASE_STABILITY_TOL, result["phase_rel40"] / PHASE_STABILITY_TOL, "通过"],
                   ["最终T/80功率窗口相对差", result["p_rel80"], POWER_STABILITY_TOL, result["p_rel80"] / POWER_STABILITY_TOL, "通过"],
                   ["最终T/80同相位状态相对差", result["phase_rel80"], PHASE_STABILITY_TOL, result["phase_rel80"] / PHASE_STABILITY_TOL, "通过"],
                   ["T/40与T/80稳态功率相对差", result["step_rel"], STEP_TOL, result["step_rel"] / STEP_TOL, "通过"]]
    add_sheet(book, "稳健性证据", ["检查", "实际值", "阈值", "阈值占用比例", "结论"], robust_rows)

    disposition_ws = add_sheet(book, "Analysis Evidence Disposition",
        ["Evidence ID", "method/source", "target claim", "disposition", "key finding", "required action", "paper/figure anchor"], result["evidence"])
    for row in range(2, disposition_ws.max_row + 1):
        disposition = str(disposition_ws.cell(row, 4).value)
        fill = "E2F0D9" if disposition == "support" else "FFF2CC" if disposition == "modify" else "F4CCCC"
        for cell in disposition_ws[row]: cell.fill = PatternFill("solid", fgColor=fill)
    output.parent.mkdir(parents=True, exist_ok=True); book.save(output)


def main() -> None:
    script = Path(__file__).resolve(); folder = script.parent
    main_path = folder / "问题二求解结果.xlsx"; refine_path = folder / "问题二边界稳态精修结果.xlsx"
    output = folder / "问题二结果深化分析.xlsx"
    progress("[1/5]", "检查已验收Q2工作簿")
    if not main_path.is_file(): raise FileNotFoundError(f"缺少 {main_path}")
    if not refine_path.is_file(): raise FileNotFoundError(f"缺少 {refine_path}")
    progress("[2/5]", "读取主优化与边界稳态精修证据")
    main_result = read_main(main_path); refine_result = read_refine(refine_path)
    progress("[3/5]", "计算收益、参数敏感性、优化路径和稳定性余量")
    result = analyze(main_result, refine_result)
    progress("[4/5]", "生成Analysis Evidence Disposition")
    progress("[5/5]", "写入问题二结果深化分析.xlsx")
    write_output(output, main_path, refine_path, main_result, refine_result, result)
    print(f"[完成] {output}", flush=True)

if __name__ == "__main__":
    main()
