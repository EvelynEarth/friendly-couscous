#!/usr/bin/env python3
"""2022A 问题二：幂律阻尼上边界 a=100000 的周期稳态一维精修。

用途：
1. 不重新运行昂贵的二维 differential_evolution；
2. 固定题面上边界 a=100000，只对 n∈[0,1] 做周期稳态功率精修；
3. 优先读取同目录 q2_result_checkpoint.json 作为已完成 Q2 主求解的候选信息；
4. 用 DOP853 + 周期同相位状态/功率窗口收敛判据直接评价稳态目标；
5. 对最终 n* 用 T/40 与 T/80 两套步长从零初值独立复算；
6. 输出 问题二边界稳态精修结果.xlsx 与 q2_boundary_refine_checkpoint.json。

本脚本只做已批准 Q2 模型的数值验证/精修，不改变模型语义。
Q2 semantic revision=1
hash=513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862
"""
from __future__ import annotations

import hashlib
import json
import math
import platform
import time
from copy import copy
from pathlib import Path

import numpy as np
import scipy
from openpyxl import Workbook, load_workbook
from scipy.integrate import solve_ivp
from scipy.optimize import minimize_scalar

SEMANTIC_REVISION = 1
SEMANTIC_HASH = "513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862"
EXPECTED_HASHES = {
    "附件3.xlsx": "50a5dd70f04dfb0a57fb2602422dc7999b30aad54ddc02353f5b8f01423fd612",
    "附件4.xlsx": "c8eff812f5980d955b4f0e587c5f7a357b2571d8d903fcb4913fba77c7354d6d",
}
LOCKED_DATA_SHA256 = "0cc51ac30576d2c2d3901aaf27dfe26526bd3768035c2685e9a57d4f81551653"
A_FIXED = 100000.0
RTOL = 1.0e-10
ATOL = 1.0e-12
POWER_STABILITY_TOL = 1.0e-5
PHASE_STABILITY_TOL = 1.0e-5
STEP_REFINEMENT_TOL = 1.0e-6
FINAL_PRE_CYCLES = 80
BLOCK_CYCLES = 10
FINAL_MAX_CYCLES = 400
LOCAL_LEFT = 0.36
LOCAL_RIGHT = 0.47
LOCAL_STEP = 0.01
SCALAR_XATOL = 2.0e-6
SCALAR_MAXITER = 24
FALLBACK_N0 = 0.4136816549748946
FALLBACK_OLD_POWER = 229.9924378779069
CHECKPOINT_SIGNATURE = (
    "q2-boundary-refine-r1|a=100000|steady-objective|DOP853|rtol=1e-10|atol=1e-12|"
    "local=0.36:0.01:0.47|final=40/80|max=400T"
)


def progress(stage: str, message: str) -> None:
    print(f"[{time.strftime('%H:%M:%S')}] {stage} {message}", flush=True)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def aggregate_data_hash(actual: dict[str, str]) -> str:
    text = "".join(f"{name}:{actual[name]}\n" for name in sorted(actual))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def atomic_write_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)


def load_json(path: Path) -> dict[str, object] | None:
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def audit_inputs(project_root: Path) -> list[list[object]]:
    rows: list[list[object]] = []
    actual: dict[str, str] = {}
    for name, expected in EXPECTED_HASHES.items():
        path = project_root / name
        if not path.is_file():
            raise FileNotFoundError(
                f"缺少输入文件: {path}。附件3.xlsx、附件4.xlsx必须直接位于A题根目录。"
            )
        current = sha256_file(path)
        actual[name] = current
        passed = current.lower() == expected.lower()
        rows.append(["通过" if passed else "阻断", f"{name} SHA-256", current, expected])
        if not passed:
            raise ValueError(f"{name} 与2022A锁定附件哈希不一致")
    agg = aggregate_data_hash(actual)
    rows.append(["通过" if agg == LOCKED_DATA_SHA256 else "阻断", "data_sha256", agg, LOCKED_DATA_SHA256])
    if agg != LOCKED_DATA_SHA256:
        raise ValueError("聚合 data_sha256 与 Q2 锁定数据不一致")
    return rows


def read_q2_parameters(project_root: Path) -> dict[str, float]:
    wb3 = load_workbook(project_root / "附件3.xlsx", data_only=True, read_only=True)
    ws3 = wb3.active
    q2_row = next((row for row in ws3.iter_rows(min_row=2, values_only=True) if row[0] == "问题2"), None)
    if q2_row is None:
        raise ValueError("附件3中未找到问题2参数行")

    wb4 = load_workbook(project_root / "附件4.xlsx", data_only=True, read_only=True)
    ws4 = wb4.active
    p4 = {
        str(row[0]): float(row[1])
        for row in ws4.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    needed = [
        "浮子质量 (kg)", "浮子底半径 (m)", "振子质量 (kg)",
        "海水的密度 (kg/m3)", "重力加速度 (m/s2)", "弹簧刚度 (N/m)",
    ]
    missing = [name for name in needed if name not in p4]
    if missing:
        raise ValueError(f"附件4缺少字段: {missing}")

    params = {
        "omega": float(q2_row[1]),
        "added_mass": float(q2_row[2]),
        "wave_damping": float(q2_row[4]),
        "force_amp": float(q2_row[6]),
        "float_mass": p4["浮子质量 (kg)"],
        "radius": p4["浮子底半径 (m)"],
        "osc_mass": p4["振子质量 (kg)"],
        "rho": p4["海水的密度 (kg/m3)"],
        "g": p4["重力加速度 (m/s2)"],
        "spring_k": p4["弹簧刚度 (N/m)"],
    }
    if not math.isclose(params["omega"], 2.2143, rel_tol=0.0, abs_tol=1e-12):
        raise ValueError("问题2频率不是锁定值 2.2143 s^-1")
    if not all(math.isfinite(v) for v in params.values()):
        raise ValueError("Q2参数含 NaN/Inf")
    return params


def hydro_stiffness(params: dict[str, float]) -> float:
    return params["rho"] * params["g"] * math.pi * params["radius"] ** 2


def state_rhs(params: dict[str, float], a_value: float, n_value: float):
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    kh = hydro_stiffness(params)
    b = params["wave_damping"]
    force = params["force_amp"]
    omega = params["omega"]

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        zf, vf, zo, vo = y
        vr = vf - vo
        fd = a_value * abs(vr) ** n_value * vr
        af = (force * math.cos(omega * t) - b * vf - kh * zf - k * (zf - zo) - fd) / mass
        ao = (k * (zf - zo) + fd) / mo
        return np.array([vf, af, vo, ao], dtype=float)

    return rhs


def integrate_state(
    params: dict[str, float], a_value: float, n_value: float,
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> np.ndarray:
    sol = solve_ivp(
        state_rhs(params, a_value, n_value), (t0, t1), y0,
        method="DOP853", rtol=RTOL, atol=ATOL, max_step=max_step,
    )
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"状态积分失败: {sol.message}")
    return np.asarray(sol.y[:, -1], dtype=float)


def integrate_power_window(
    params: dict[str, float], a_value: float, n_value: float,
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> tuple[np.ndarray, float]:
    base = state_rhs(params, a_value, n_value)

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        dy = base(t, y[:4])
        vr = float(y[1] - y[3])
        power = a_value * abs(vr) ** (n_value + 2.0)
        return np.array([*dy, power], dtype=float)

    initial = np.array([*y0, 0.0], dtype=float)
    sol = solve_ivp(
        rhs, (t0, t1), initial, method="DOP853",
        rtol=RTOL, atol=ATOL, max_step=max_step,
    )
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"功率积分失败: {sol.message}")
    avg = float(sol.y[4, -1] / (t1 - t0))
    if not math.isfinite(avg) or avg < -1e-10:
        raise RuntimeError("平均功率出现非有限值或明显负值")
    return np.asarray(sol.y[:4, -1], dtype=float), max(0.0, avg)


def converge_from_state(
    params: dict[str, float], n_value: float, y0: np.ndarray,
    step_divisor: float, max_blocks: int = 32,
) -> dict[str, object]:
    """From a same-phase seed, iterate 10T blocks until power and state both converge."""
    T = 2.0 * math.pi / params["omega"]
    max_step = T / step_divisor
    y = np.asarray(y0, dtype=float).copy()
    previous_end: np.ndarray | None = None
    previous_power: float | None = None
    diagnostics: list[list[object]] = []

    for block in range(1, max_blocks + 1):
        t0 = (block - 1) * BLOCK_CYCLES * T
        t1 = block * BLOCK_CYCLES * T
        end, power = integrate_power_window(params, A_FIXED, n_value, t0, t1, y, max_step)
        if previous_end is not None and previous_power is not None:
            power_rel = abs(power - previous_power) / (1.0 + abs(power))
            phase_rel = float(np.linalg.norm(end - previous_end) / (1.0 + np.linalg.norm(end)))
            passed = power_rel <= POWER_STABILITY_TOL and phase_rel <= PHASE_STABILITY_TOL
            diagnostics.append([
                block - 1, block, power, power_rel, phase_rel, passed,
            ])
            if passed:
                return {
                    "n": float(n_value), "power": float(power), "end_state": end,
                    "extra_cycles": block * BLOCK_CYCLES, "diagnostics": diagnostics,
                }
        previous_end = end.copy()
        previous_power = power
        y = end

    last = diagnostics[-1] if diagnostics else [None, None, None, math.inf, math.inf, False]
    raise RuntimeError(
        f"n={n_value:.10f} 从warm-start出发在{max_blocks * BLOCK_CYCLES}T内仍未稳定；"
        f"最后功率相对差={last[3]:.3e}, 同相位状态相对差={last[4]:.3e}"
    )


def verify_from_zero(params: dict[str, float], n_value: float, step_divisor: float) -> dict[str, object]:
    """Independent full verification from zero initial conditions."""
    T = 2.0 * math.pi / params["omega"]
    max_step = T / step_divisor
    y = integrate_state(
        params, A_FIXED, n_value, 0.0, FINAL_PRE_CYCLES * T,
        np.zeros(4, dtype=float), max_step,
    )
    previous_end: np.ndarray | None = None
    previous_power: float | None = None
    diagnostics: list[list[object]] = []
    cycle = FINAL_PRE_CYCLES
    while cycle < FINAL_MAX_CYCLES:
        nxt = cycle + BLOCK_CYCLES
        end, power = integrate_power_window(params, A_FIXED, n_value, cycle * T, nxt * T, y, max_step)
        if previous_end is not None and previous_power is not None:
            power_rel = abs(power - previous_power) / (1.0 + abs(power))
            phase_rel = float(np.linalg.norm(end - previous_end) / (1.0 + np.linalg.norm(end)))
            passed = power_rel <= POWER_STABILITY_TOL and phase_rel <= PHASE_STABILITY_TOL
            diagnostics.append([cycle - BLOCK_CYCLES, cycle, nxt, previous_power, power, power_rel, phase_rel, passed])
            if passed:
                return {
                    "n": float(n_value), "power": float(power), "end_state": end,
                    "stop_cycle": int(nxt), "diagnostics": diagnostics,
                }
        previous_end = end.copy()
        previous_power = power
        y = end
        cycle = nxt
    raise RuntimeError(f"n={n_value:.10f} 从零初值验证到{FINAL_MAX_CYCLES}T仍未满足稳态阈值")


class SteadyEvaluator:
    def __init__(self, params: dict[str, float], checkpoint_path: Path):
        self.params = params
        self.checkpoint_path = checkpoint_path
        self.cache: dict[float, dict[str, object]] = {}
        self.eval_count = 0
        self._load_checkpoint()

    @staticmethod
    def key(n_value: float) -> float:
        return round(float(n_value), 10)

    def _load_checkpoint(self) -> None:
        saved = load_json(self.checkpoint_path)
        if not saved:
            return
        if saved.get("semantic_hash") != SEMANTIC_HASH:
            return
        if saved.get("data_sha256") != LOCKED_DATA_SHA256:
            return
        if saved.get("signature") != CHECKPOINT_SIGNATURE:
            return
        entries = saved.get("cache", [])
        if not isinstance(entries, list):
            return
        for item in entries:
            if not isinstance(item, dict):
                continue
            try:
                n_value = self.key(float(item["n"]))
                power = float(item["power"])
                state = np.asarray(item["end_state"], dtype=float)
            except (KeyError, TypeError, ValueError):
                continue
            if state.shape == (4,) and np.all(np.isfinite(state)) and math.isfinite(power):
                self.cache[n_value] = {
                    "n": n_value, "power": power, "end_state": state,
                    "extra_cycles": int(item.get("extra_cycles", 0)), "source": "checkpoint",
                }
        if self.cache:
            progress("[恢复]", f"载入边界精修 checkpoint：{len(self.cache)} 个稳态评价点")

    def save(self) -> None:
        payload = {
            "semantic_hash": SEMANTIC_HASH,
            "data_sha256": LOCKED_DATA_SHA256,
            "signature": CHECKPOINT_SIGNATURE,
            "a_fixed": A_FIXED,
            "cache": [
                {
                    "n": float(item["n"]),
                    "power": float(item["power"]),
                    "end_state": np.asarray(item["end_state"], dtype=float).tolist(),
                    "extra_cycles": int(item.get("extra_cycles", 0)),
                }
                for _, item in sorted(self.cache.items())
            ],
        }
        atomic_write_json(self.checkpoint_path, payload)

    def seed(self, n_value: float) -> dict[str, object]:
        key = self.key(n_value)
        if key in self.cache:
            return self.cache[key]
        progress("[3/7]", f"建立中心稳态种子 n={n_value:.8f}（从零初值）")
        result = verify_from_zero(self.params, n_value, 40.0)
        item = {
            "n": key, "power": float(result["power"]),
            "end_state": np.asarray(result["end_state"], dtype=float),
            "extra_cycles": int(result["stop_cycle"]), "source": "zero-seed",
        }
        self.cache[key] = item
        self.save()
        progress("[3/7]", f"中心种子完成：P={item['power']:.9f} W, stop={item['extra_cycles']}T")
        return item

    def evaluate(self, n_value: float) -> float:
        n_value = min(1.0, max(0.0, float(n_value)))
        key = self.key(n_value)
        if key in self.cache:
            return float(self.cache[key]["power"])
        if not self.cache:
            self.seed(n_value)
            return float(self.cache[key]["power"])

        nearest_key = min(self.cache, key=lambda existing: abs(existing - key))
        seed_state = np.asarray(self.cache[nearest_key]["end_state"], dtype=float)
        self.eval_count += 1
        progress(
            "[评价]",
            f"#{self.eval_count}: n={n_value:.8f}, warm-start from n={nearest_key:.8f}",
        )
        result = converge_from_state(self.params, n_value, seed_state, 40.0)
        item = {
            "n": key, "power": float(result["power"]),
            "end_state": np.asarray(result["end_state"], dtype=float),
            "extra_cycles": int(result["extra_cycles"]), "source": "warm-start",
        }
        self.cache[key] = item
        self.save()
        progress(
            "[评价]",
            f"n={n_value:.8f} -> P={item['power']:.9f} W, additional={item['extra_cycles']}T",
        )
        return float(item["power"])


def candidate_from_previous(result_checkpoint: Path) -> tuple[float, float, str]:
    data = load_json(result_checkpoint)
    if (
        data
        and data.get("semantic_hash") == SEMANTIC_HASH
        and data.get("data_sha256") == LOCKED_DATA_SHA256
        and isinstance(data.get("nonlinear"), dict)
    ):
        nonlinear = data["nonlinear"]
        try:
            n0 = float(nonlinear["n"])
            old_power = float(nonlinear["final_power"])
            return n0, old_power, "q2_result_checkpoint.json"
        except (KeyError, TypeError, ValueError):
            pass
    return FALLBACK_N0, FALLBACK_OLD_POWER, "内置已验收候选回退值"


def excel_scalar(value: object) -> object:
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (list, tuple, dict, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def add_sheet(book: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = book.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append([excel_scalar(v) for v in row])
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column in ws.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(48, max(11, width))


def write_workbook(
    output: Path, code_hash: str, n0: float, old_power: float, source: str,
    local_rows: list[list[object]], scalar_rows: list[list[object]],
    final40: dict[str, object], final80: dict[str, object],
    boundary_rows: list[list[object]], quality_rows: list[list[object]],
) -> None:
    book = Workbook()
    book.remove(book.active)
    runtime_rows = [
        ["semantic_revision", SEMANTIC_REVISION],
        ["semantic_hash", SEMANTIC_HASH],
        ["data_sha256", LOCKED_DATA_SHA256],
        ["task", "固定a=100000，对n做周期稳态边界精修"],
        ["solver", "DOP853 + warm-start steady scan + bounded minimize_scalar + zero-start T/40/T/80 verification"],
        ["rtol", RTOL], ["atol", ATOL], ["a_fixed", A_FIXED],
        ["local_interval", f"[{LOCAL_LEFT}, {LOCAL_RIGHT}]"],
        ["local_step", LOCAL_STEP], ["scalar_xatol", SCALAR_XATOL],
        ["previous_candidate_source", source], ["previous_n", n0], ["previous_final_power_W", old_power],
        ["code_sha256", code_hash], ["scipy_version", scipy.__version__], ["platform", platform.platform()],
    ]
    add_sheet(book, "运行配置", ["项目", "值"], runtime_rows)

    best_power = float(final40["power"])
    best_n = float(final40["n"])
    core_rows = [
        ["固定比例系数a", A_FIXED, "随n变化的题设比例系数"],
        ["边界精修最优n", best_n, "1"],
        ["边界精修最大稳态平均功率(T/40)", best_power, "W"],
        ["步长加密平均功率(T/80)", float(final80["power"]), "W"],
        ["相对原二维候选功率增量", best_power - old_power, "W"],
        ["相对原二维候选提升", (best_power - old_power) / old_power if old_power else math.nan, "1"],
        ["T/40最终停止周期", int(final40["stop_cycle"]), "T"],
        ["T/80最终停止周期", int(final80["stop_cycle"]), "T"],
    ]
    add_sheet(book, "核心结论", ["指标", "数值", "单位/说明"], core_rows)
    add_sheet(book, "局部稳态扫描", ["n", "稳态平均功率/W", "备注"], local_rows)
    add_sheet(book, "一维精修", ["项目", "数值", "说明"], scalar_rows)
    add_sheet(book, "边界对照", ["n", "稳态平均功率/W", "相对精修最优差/W", "用途"], boundary_rows)

    convergence_rows: list[list[object]] = []
    for label, result in [("最终T/40", final40), ("最终T/80", final80)]:
        for row in result["diagnostics"]:
            convergence_rows.append([label, *row])
    add_sheet(
        book, "收敛诊断",
        ["对象", "前窗口起始周期", "前窗口结束周期", "后窗口结束周期", "前功率/W", "后功率/W", "功率相对差", "同相位状态相对差", "通过"],
        convergence_rows,
    )
    add_sheet(book, "主结果质量门", ["检查项", "是否通过", "证据"], quality_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)


def main() -> None:
    script_path = Path(__file__).resolve()
    project_root = script_path.parents[1]
    work_dir = script_path.parent
    result_checkpoint = work_dir / "q2_result_checkpoint.json"
    refine_checkpoint = work_dir / "q2_boundary_refine_checkpoint.json"
    output = work_dir / "问题二边界稳态精修结果.xlsx"

    progress("[1/7]", "校验附件3/附件4与Q2数据作用域")
    audit_inputs(project_root)
    params = read_q2_parameters(project_root)

    progress("[2/7]", "读取已有Q2候选，避免重新二维全局搜索")
    n0, old_power, source = candidate_from_previous(result_checkpoint)
    progress("[2/7]", f"已有候选：n={n0:.10f}, P={old_power:.9f} W, source={source}")

    evaluator = SteadyEvaluator(params, refine_checkpoint)
    evaluator.seed(n0)

    progress("[4/7]", f"固定a={A_FIXED:.0f}，执行局部稳态扫描 {LOCAL_LEFT:.2f}~{LOCAL_RIGHT:.2f}")
    grid = np.arange(LOCAL_LEFT, LOCAL_RIGHT + 0.5 * LOCAL_STEP, LOCAL_STEP)
    scan_points = sorted({round(float(x), 10) for x in grid} | {round(n0, 10)}, key=lambda x: abs(x - n0))
    for idx, n_value in enumerate(scan_points, start=1):
        power = evaluator.evaluate(n_value)
        progress("[4/7]", f"扫描 {idx}/{len(scan_points)}: n={n_value:.8f}, P={power:.9f} W")

    local_sorted = sorted((n, float(item["power"])) for n, item in evaluator.cache.items() if LOCAL_LEFT - 1e-12 <= n <= LOCAL_RIGHT + 1e-12)
    if len(local_sorted) < 3:
        raise RuntimeError("局部扫描有效点不足，无法构造一维精修区间")
    best_index = max(range(len(local_sorted)), key=lambda i: local_sorted[i][1])
    if best_index == 0 or best_index == len(local_sorted) - 1:
        raise RuntimeError("局部扫描最大值落在扫描区间边界，请扩大LOCAL_LEFT/LOCAL_RIGHT后重跑")
    bracket_left = local_sorted[best_index - 1][0]
    bracket_mid = local_sorted[best_index][0]
    bracket_right = local_sorted[best_index + 1][0]
    progress(
        "[5/7]",
        f"稳态一维精修区间 [{bracket_left:.8f}, {bracket_right:.8f}]，局部最佳 n={bracket_mid:.8f}",
    )

    scalar_calls: list[list[object]] = []

    def objective(n_value: float) -> float:
        power = evaluator.evaluate(float(n_value))
        scalar_calls.append([len(scalar_calls) + 1, float(n_value), power])
        return -power

    scalar = minimize_scalar(
        objective,
        bounds=(bracket_left, bracket_right),
        method="bounded",
        options={"xatol": SCALAR_XATOL, "maxiter": SCALAR_MAXITER},
    )
    if not scalar.success or not math.isfinite(float(scalar.x)) or not math.isfinite(float(scalar.fun)):
        raise RuntimeError(f"一维稳态精修未正常终止: {scalar.message}")
    n_star = float(scalar.x)
    p_warm = -float(scalar.fun)
    progress("[5/7]", f"一维精修完成：n≈{n_star:.10f}, warm-start P≈{p_warm:.9f} W")

    progress("[6/7]", "最终候选从零初值独立验证：T/40")
    final40 = verify_from_zero(params, n_star, 40.0)
    progress("[6/7]", f"T/40: P={float(final40['power']):.9f} W, stop={final40['stop_cycle']}T")
    progress("[6/7]", "步长加密独立验证：T/80")
    final80 = verify_from_zero(params, n_star, 80.0)
    progress("[6/7]", f"T/80: P={float(final80['power']):.9f} W, stop={final80['stop_cycle']}T")

    step_error = abs(float(final80["power"]) - float(final40["power"])) / (1.0 + abs(float(final80["power"])))
    if step_error > STEP_REFINEMENT_TOL:
        raise RuntimeError(f"边界精修最终候选步长加密未通过: {step_error:.3e}")

    progress("[6/7]", "检查固定a=100000时 n=0 与 n=1 边界")
    p_n0 = evaluator.evaluate(0.0)
    p_n1 = evaluator.evaluate(1.0)
    p_final = float(final40["power"])
    if p_n0 > p_final + 1e-5 or p_n1 > p_final + 1e-5:
        raise RuntimeError("固定a=100000的一维n边界出现高于精修候选的稳态功率")

    local_rows = [
        [n, power, "局部稳态扫描" if abs(n - n0) > 1e-9 else "原二维候选n（改用a=100000稳态复算）"]
        for n, power in sorted(local_sorted)
    ]
    scalar_rows = [
        ["优化方法", "bounded minimize_scalar", "目标直接为周期稳态平均功率"],
        ["精修左端", bracket_left, "n"],
        ["局部扫描最佳", bracket_mid, "n"],
        ["精修右端", bracket_right, "n"],
        ["精修n*", n_star, "n"],
        ["warm-start目标功率", p_warm, "W"],
        ["success", bool(scalar.success), str(scalar.message)],
        ["nfev", int(getattr(scalar, "nfev", len(scalar_calls))), "稳态目标调用次数"],
    ]
    for call_id, n_value, power in scalar_calls:
        scalar_rows.append([f"call_{call_id:02d}", n_value, f"P={power:.9f} W"])

    boundary_rows = [
        [0.0, p_n0, p_n0 - p_final, "固定a=100000的n下边界"],
        [1.0, p_n1, p_n1 - p_final, "固定a=100000的n上边界"],
        [n0, evaluator.evaluate(n0), evaluator.evaluate(n0) - p_final, "原二维候选n在a=100000下的稳态复算"],
        [n_star, p_final, 0.0, "边界稳态精修最终候选"],
    ]

    improved = p_final >= old_power - 1e-5
    quality_rows = [
        ["Q2语义与数据作用域", "是", f"semantic_hash={SEMANTIC_HASH}; data_sha256={LOCKED_DATA_SHA256}"],
        ["比例系数边界", "是", "固定a=100000，严格位于题面允许上边界"],
        ["稳态目标精修", "是", f"局部扫描+bounded一维精修，n*={n_star:.10f}"],
        ["最终周期稳态", "是", f"T/40从零初值验证至{final40['stop_cycle']}T通过功率/同相位状态阈值"],
        ["步长加密", "是" if step_error <= STEP_REFINEMENT_TOL else "否", f"相对功率差={step_error:.3e} <= {STEP_REFINEMENT_TOL:.1e}"],
        ["n边界检查", "是", f"P(n=0)={p_n0:.9f} W, P(n=1)={p_n1:.9f} W 均不高于精修候选"],
        ["相对原二维候选", "是" if improved else "警告", f"新稳态P={p_final:.9f} W, 原候选P={old_power:.9f} W, 差={p_final-old_power:+.9f} W"],
    ]

    progress("[7/7]", "写入边界稳态精修工作簿")
    write_workbook(
        output, sha256_file(script_path), n0, old_power, source,
        local_rows, scalar_rows, final40, final80, boundary_rows, quality_rows,
    )
    evaluator.save()
    progress("[完成]", f"输出：{output}")
    progress("[结论]", f"a*=100000, n*≈{n_star:.10f}, P≈{p_final:.9f} W, step_error={step_error:.3e}")


if __name__ == "__main__":
    main()
