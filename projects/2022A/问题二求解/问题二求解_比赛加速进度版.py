#!/usr/bin/env python3
"""2022A 问题二：PTO 阻尼参数与平均输出功率优化主求解。

本脚本必须由用户本地 full_fidelity 执行。它只读取项目根目录中的附件3、附件4，
不读取网上答案，也不沿用问题一的水动力数值。输出：
    问题二求解/问题二求解结果.xlsx

模型语义绑定：
Q2 semantic revision=1
hash=513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862
"""
from __future__ import annotations

import hashlib
import json
import math
import os
import platform
import time
from multiprocessing import freeze_support
from copy import copy
from pathlib import Path

import numpy as np
import scipy
from openpyxl import Workbook, load_workbook
from scipy.integrate import solve_ivp
from scipy.optimize import differential_evolution, minimize


SEMANTIC_REVISION = 1
SEMANTIC_HASH = "513d81ccab68ccea4f5db5df0ff8b87169240d27c55f1d4edef80e67606f7862"
EXPECTED_HASHES = {
    "附件3.xlsx": "50a5dd70f04dfb0a57fb2602422dc7999b30aad54ddc02353f5b8f01423fd612",
    "附件4.xlsx": "c8eff812f5980d955b4f0e587c5f7a357b2571d8d903fcb4913fba77c7354d6d",
}
LOCKED_DATA_SHA256 = "0cc51ac30576d2c2d3901aaf27dfe26526bd3768035c2685e9a57d4f81551653"
C_MAX = 100000.0
A_MAX = 100000.0
RTOL = 1.0e-10
ATOL = 1.0e-12
DE_SEED = 2022
DE_MAXITER = 30
DE_POPSIZE = 12
DE_TOL = 1.0e-5
SEARCH_TRANSIENT_CYCLES = 50
SEARCH_AVERAGE_CYCLES = 10
FINAL_PRE_CYCLES = 80
FINAL_BLOCK_CYCLES = 10
FINAL_MAX_CYCLES = 400
POWER_STABILITY_TOL = 1.0e-5
PHASE_STABILITY_TOL = 1.0e-5
LINEAR_CROSS_TOL = 1.0e-6
N0_EQUIV_TOL = 1.0e-8
STEP_REFINEMENT_TOL = 1.0e-6
NEIGHBOR_TOL = 2.0e-4

# 比赛运行工程配置：不改变模型、搜索域、ODE精度或质量门，只改善可观测性与计算利用率。
DEFAULT_WORKERS = min(8, max(1, (os.cpu_count() or 2) - 2))
MAX_WORKERS = max(1, int(os.environ.get("Q2_WORKERS", DEFAULT_WORKERS)))
COMPUTE_SIGNATURE = (
    "q2-r1|DOP853|rtol=1e-10|atol=1e-12|DE30x12|search=50T+10T|"
    "final_pre=80T|block=10T|max=400T|step=40/80|neighbor=2e-4"
)
POWER_CACHE: dict[tuple[float, float], float] = {}

FULL_FIDELITY_CONFIG = {
    "execution_owner": "user",
    "execution_profile": "full_fidelity",
    "stage": "primary",
    "problem_name": "问题二",
    "semantic_revision": SEMANTIC_REVISION,
    "semantic_hash": SEMANTIC_HASH,
    "data_paths": ["附件3.xlsx", "附件4.xlsx"],
    "data_sha256": LOCKED_DATA_SHA256,
    "solver": "DOP853 + analytic linear optimum + parallel differential_evolution + bounded Powell refinement",
    "solver_version": "SciPy runtime version is recorded in workbook",
    "random_seed": DE_SEED,
    "parallel_workers": MAX_WORKERS,
    "compute_signature": COMPUTE_SIGNATURE,
    "tolerance": RTOL,
    "iteration_or_time_limit": (
        "DE maxiter=30,popsize=12; each nonlinear search candidate uses 60T "
        "(last 10T power); final verification >=100T and extends to 400T if needed"
    ),
    "expected_workbook": "问题二求解/问题二求解结果.xlsx",
    "de_checkpoint": "问题二求解/q2_de_checkpoint.json",
    "result_checkpoint": "问题二求解/q2_result_checkpoint.json",
    "allow_reduced_data": False,
    "allow_coarser_grid": False,
    "allow_shorter_horizon": False,
    "allow_fewer_repetitions": False,
    "allow_relaxed_tolerance": False,
    "allow_silent_solver_fallback": False,
}


def progress(stage: str, message: str) -> None:
    """Print timestamped progress immediately so PyCharm never looks frozen."""
    stamp = time.strftime("%H:%M:%S")
    print(f"[{stamp}] {stage} {message}", flush=True)


def atomic_write_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    tmp.replace(path)


def read_json_if_valid(path: Path) -> dict[str, object] | None:
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if data.get("semantic_hash") != SEMANTIC_HASH:
        return None
    if data.get("data_sha256") != LOCKED_DATA_SHA256:
        return None
    if data.get("compute_signature") != COMPUTE_SIGNATURE:
        return None
    return data


def jsonable(value: object) -> object:
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [jsonable(v) for v in value]
    return value


class SearchObjective:
    """Pickle-safe DE objective for Windows multiprocessing."""

    def __init__(self, params: dict[str, float]):
        self.params = params

    def __call__(self, x: np.ndarray) -> float:
        u = min(1.0, max(0.0, float(x[0])))
        n = min(1.0, max(0.0, float(x[1])))
        return -search_power(self.params, A_MAX * u, n)


def cached_search_power(params: dict[str, float], a_value: float, n_value: float) -> float:
    """Cache sequential Powell/neighborhood evaluations; DE workers remain independent."""
    key = (round(float(a_value), 8), round(float(n_value), 10))
    if key not in POWER_CACHE:
        POWER_CACHE[key] = search_power(params, float(a_value), float(n_value))
    return POWER_CACHE[key]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def aggregate_data_hash(actual_hashes: dict[str, str]) -> str:
    text = "".join(f"{name}:{actual_hashes[name]}\n" for name in sorted(actual_hashes))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def audit_source_files(project_root: Path) -> tuple[dict[str, str], list[list[object]]]:
    actual: dict[str, str] = {}
    rows: list[list[object]] = []
    for name, expected in EXPECTED_HASHES.items():
        path = project_root / name
        if not path.is_file():
            raise FileNotFoundError(
                f"缺少输入文件: {path}。请把附件3.xlsx、附件4.xlsx直接放在A题根目录，"
                "与‘问题二求解’文件夹同级。"
            )
        current = sha256_file(path)
        actual[name] = current
        passed = current.lower() == expected.lower()
        rows.append(["通过" if passed else "阻断", f"{name} SHA-256", current, expected])
        if not passed:
            raise ValueError(f"{name} 哈希与2022A锁定附件不一致，停止求解")
    if aggregate_data_hash(actual) != LOCKED_DATA_SHA256:
        raise ValueError("Q2聚合data_sha256与锁定配置不一致")
    return actual, rows


def read_q2_parameters(project_root: Path) -> tuple[dict[str, float], list[list[object]]]:
    wb3 = load_workbook(project_root / "附件3.xlsx", data_only=True, read_only=True)
    ws3 = wb3.active
    expected3 = [
        "问题", "入射波浪频率 (s-1)", "垂荡附加质量 (kg)", "纵摇附加转动惯量 (kg·m2)",
        "垂荡兴波阻尼系数 (N·s/m)", "纵摇兴波阻尼系数 (N·m·s)",
        "垂荡激励力振幅 (N)", "纵摇激励力矩振幅 (N·m)",
    ]
    headers3 = [ws3.cell(1, col).value for col in range(1, 9)]
    if headers3 != expected3:
        raise ValueError("附件3表头与2022A锁定版本不一致")
    q2_row = next((row for row in ws3.iter_rows(min_row=2, values_only=True) if row[0] == "问题2"), None)
    if q2_row is None:
        raise ValueError("附件3中未找到问题2参数行")

    wb4 = load_workbook(project_root / "附件4.xlsx", data_only=True, read_only=True)
    ws4 = wb4.active
    if [ws4["A1"].value, ws4["B1"].value] != ["参数", "取值"]:
        raise ValueError("附件4表头与2022A锁定版本不一致")
    p4 = {str(row[0]): float(row[1]) for row in ws4.iter_rows(min_row=2, values_only=True) if row[0] is not None}
    needed = ["浮子质量 (kg)", "浮子底半径 (m)", "振子质量 (kg)", "海水的密度 (kg/m3)",
              "重力加速度 (m/s2)", "弹簧刚度 (N/m)"]
    if any(name not in p4 for name in needed):
        raise ValueError("附件4缺少Q2所需字段")

    params = {
        "omega": float(q2_row[1]),
        "added_mass": float(q2_row[2]),
        "wave_damping": float(q2_row[4]),
        "force_amp": float(q2_row[6]),
        "float_mass": p4["浮子质量 (kg)"],
        "radius": p4["浮子底半径 (m)"],
        "osc_mass": p4["振子质量 (kg)"],
        "rho": p4["海水的密度 (kg/m3)"],
        "g": p4["重力加速度 (m/s2)"],
        "spring_k": p4["弹簧刚度 (N/m)"],
    }
    if not all(math.isfinite(value) for value in params.values()):
        raise ValueError("Q2参数存在NaN或Inf")
    if not math.isclose(params["omega"], 2.2143, rel_tol=0.0, abs_tol=1.0e-12):
        raise ValueError("附件3问题2入射波浪频率不是锁定值2.2143 s^-1")
    rows = [
        ["通过", "附件3问题2参数", "ω、垂荡附加质量、垂荡兴波阻尼、激励力均直接读取问题2行", "不沿用Q1数值"],
        ["通过", "附件4字段", "Q2所需质量/半径/密度/g/弹簧刚度齐全", "原始值直接使用"],
        ["通过", "预处理判定", "not_needed", "仅非破坏性字段/单位/NaN/Inf审计"],
    ]
    return params, rows


def hydro_stiffness(params: dict[str, float]) -> float:
    return params["rho"] * params["g"] * math.pi * params["radius"] ** 2


def damper_force(vr: float, damping: tuple[str, float, float]) -> float:
    mode, p1, p2 = damping
    if mode == "linear":
        return p1 * vr
    if mode == "power":
        return p1 * abs(vr) ** p2 * vr
    raise ValueError(f"未知阻尼模式: {mode}")


def instant_power(vr: float, damping: tuple[str, float, float]) -> float:
    mode, p1, p2 = damping
    if mode == "linear":
        return p1 * vr * vr
    if mode == "power":
        return p1 * abs(vr) ** (p2 + 2.0)
    raise ValueError(f"未知阻尼模式: {mode}")


def state_rhs(params: dict[str, float], damping: tuple[str, float, float]):
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    kh = hydro_stiffness(params)
    b = params["wave_damping"]
    force_amp = params["force_amp"]
    omega = params["omega"]

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        zf, vf, zo, vo = y
        vr = vf - vo
        fd = damper_force(vr, damping)
        af = (force_amp * math.cos(omega * t) - b * vf - kh * zf - k * (zf - zo) - fd) / mass
        ao = (k * (zf - zo) + fd) / mo
        return np.array([vf, af, vo, ao], dtype=float)

    return rhs


def integrate_state(
    params: dict[str, float], damping: tuple[str, float, float],
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> np.ndarray:
    sol = solve_ivp(
        state_rhs(params, damping), (t0, t1), y0, method="DOP853",
        rtol=RTOL, atol=ATOL, max_step=max_step,
    )
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"状态积分失败: {sol.message}")
    return np.asarray(sol.y[:, -1], dtype=float)


def integrate_power_window(
    params: dict[str, float], damping: tuple[str, float, float],
    t0: float, t1: float, y0: np.ndarray, max_step: float,
) -> tuple[np.ndarray, float]:
    base_rhs = state_rhs(params, damping)

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        dy = base_rhs(t, y[:4])
        vr = float(y[1] - y[3])
        return np.array([*dy, instant_power(vr, damping)], dtype=float)

    initial = np.array([*y0, 0.0], dtype=float)
    sol = solve_ivp(rhs, (t0, t1), initial, method="DOP853", rtol=RTOL, atol=ATOL, max_step=max_step)
    if not sol.success or not np.all(np.isfinite(sol.y[:, -1])):
        raise RuntimeError(f"功率积分失败: {sol.message}")
    end = np.asarray(sol.y[:4, -1], dtype=float)
    average_power = float(sol.y[4, -1] / (t1 - t0))
    if average_power < -1.0e-10 or not math.isfinite(average_power):
        raise RuntimeError("平均PTO功率出现非有限值或明显负值")
    return end, max(0.0, average_power)


def linear_power_coefficients(params: dict[str, float]) -> dict[str, float]:
    omega = params["omega"]
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    h = hydro_stiffness(params) - mass * omega ** 2
    bomega = params["wave_damping"] * omega
    gterm = mo * omega ** 2 - k
    a0 = h * gterm + mo * omega ** 2 * k
    b0 = bomega * gterm
    a1 = bomega * omega
    b1 = omega * (mo * omega ** 2 - h)
    alpha = a1 * a1 + b1 * b1
    beta = 2.0 * (a0 * a1 + b0 * b1)
    gamma = a0 * a0 + b0 * b0
    numerator_factor = 0.5 * omega ** 2 * params["force_amp"] ** 2 * (mo * omega ** 2) ** 2
    if alpha <= 0.0 or gamma <= 0.0:
        raise RuntimeError("直线阻尼解析功率系数退化，无法使用当前闭式最优公式")
    return {
        "H": h, "B": bomega, "G": gterm, "A0": a0, "B0": b0, "A1": a1, "B1": b1,
        "alpha": alpha, "beta": beta, "gamma": gamma, "numerator_factor": numerator_factor,
    }


def linear_power_closed_form(c_value: float, coeff: dict[str, float]) -> float:
    denominator = coeff["alpha"] * c_value ** 2 + coeff["beta"] * c_value + coeff["gamma"]
    if denominator <= 0.0:
        raise RuntimeError("直线阻尼功率解析式分母非正")
    return coeff["numerator_factor"] * c_value / denominator


def linear_optimum(params: dict[str, float]) -> tuple[float, float, dict[str, float]]:
    coeff = linear_power_coefficients(params)
    c0 = math.sqrt(coeff["gamma"] / coeff["alpha"])
    c_star = min(C_MAX, max(0.0, c0))
    p_star = linear_power_closed_form(c_star, coeff)
    endpoint_best = max(linear_power_closed_form(0.0, coeff), linear_power_closed_form(C_MAX, coeff))
    if p_star + 1.0e-10 < endpoint_best:
        raise RuntimeError("解析驻点未优于区间端点，直线阻尼最优性检查失败")
    return c_star, p_star, coeff


def search_power(params: dict[str, float], a_value: float, n_value: float) -> float:
    period = 2.0 * math.pi / params["omega"]
    damping = ("power", float(a_value), float(n_value))
    y = integrate_state(params, damping, 0.0, SEARCH_TRANSIENT_CYCLES * period, np.zeros(4), period / 20.0)
    _, power = integrate_power_window(
        params, damping, SEARCH_TRANSIENT_CYCLES * period,
        (SEARCH_TRANSIENT_CYCLES + SEARCH_AVERAGE_CYCLES) * period, y, period / 20.0,
    )
    return power


def verify_periodic_candidate(
    params: dict[str, float], damping: tuple[str, float, float], step_divisor: float,
    require_phase: bool = True, label: str = "候选",
) -> dict[str, object]:
    period = 2.0 * math.pi / params["omega"]
    max_step = period / step_divisor
    progress("[验证]", f"{label}: 先积分到 {FINAL_PRE_CYCLES}T，max_step=T/{step_divisor:g}")
    y = integrate_state(params, damping, 0.0, FINAL_PRE_CYCLES * period, np.zeros(4), max_step)
    previous_end: np.ndarray | None = None
    previous_power: float | None = None
    diagnostics: list[list[object]] = []
    last_metrics: tuple[float, float] | None = None
    cycle = FINAL_PRE_CYCLES
    while cycle < FINAL_MAX_CYCLES:
        next_cycle = cycle + FINAL_BLOCK_CYCLES
        end, power = integrate_power_window(params, damping, cycle * period, next_cycle * period, y, max_step)
        if previous_end is not None and previous_power is not None:
            power_rel = abs(power - previous_power) / (1.0 + abs(power))
            phase_rel = float(np.linalg.norm(end - previous_end) / (1.0 + np.linalg.norm(end)))
            last_metrics = (power_rel, phase_rel)
            phase_ok = (phase_rel <= PHASE_STABILITY_TOL) if require_phase else True
            passed = power_rel <= POWER_STABILITY_TOL and phase_ok
            diagnostics.append([cycle - FINAL_BLOCK_CYCLES, cycle, next_cycle, previous_power, power, power_rel, phase_rel, passed])
            progress(
                "[验证]",
                f"{label}: {cycle-FINAL_BLOCK_CYCLES:.0f}T→{next_cycle:.0f}T, "
                f"P={power:.6f} W, ΔP={power_rel:.3e}, Δphase={phase_rel:.3e}, "
                f"{'通过' if passed else '继续'}",
            )
            if passed:
                return {"power": power, "end_state": end, "stop_cycle": next_cycle, "diagnostics": diagnostics}
        else:
            progress("[验证]", f"{label}: 完成 {cycle:.0f}T→{next_cycle:.0f}T 首个比较窗口")
        previous_end = end.copy()
        previous_power = power
        y = end
        cycle = next_cycle
    criterion = "功率窗口与同相位状态" if require_phase else "功率窗口"
    if last_metrics is None:
        raise RuntimeError(f"候选点在{FINAL_MAX_CYCLES}T内未形成可比较的稳定性窗口")
    power_rel, phase_rel = last_metrics
    raise RuntimeError(
        f"候选点在{FINAL_MAX_CYCLES}T内未满足{criterion}稳定性阈值；"
        f"最后功率相对差={power_rel:.3e}，同相位状态相对差={phase_rel:.3e}"
    )


def nonlinear_optimum(
    params: dict[str, float], c_linear: float, checkpoint_path: Path,
) -> tuple[dict[str, object], list[list[object]]]:
    objective = SearchObjective(params)
    x0 = np.array([c_linear / A_MAX, 0.0], dtype=float)
    checkpoint = read_json_if_valid(checkpoint_path)
    de_payload: dict[str, object] | None = None
    init: str | np.ndarray = "latinhypercube"
    remaining = DE_MAXITER
    resumed_generation = 0

    if checkpoint is not None:
        stage = str(checkpoint.get("stage", ""))
        population = checkpoint.get("population")
        if stage == "de_complete":
            de_payload = checkpoint
            progress("[4/8]", "检测到完整 DE checkpoint，跳过全局搜索，直接进入 Powell 精修。")
        elif stage == "de_running" and isinstance(population, list) and population:
            arr = np.asarray(population, dtype=float)
            if arr.ndim == 2 and arr.shape[1] == 2 and arr.shape[0] >= 5:
                init = arr
                resumed_generation = int(checkpoint.get("generation", 0))
                remaining = max(1, DE_MAXITER - resumed_generation)
                best_x = checkpoint.get("best_x")
                if isinstance(best_x, list) and len(best_x) == 2:
                    x0 = np.asarray(best_x, dtype=float)
                progress("[4/8]", f"发现第 {resumed_generation} 代 checkpoint，从已保存种群继续，剩余最多 {remaining} 代。")

    if de_payload is None:
        generation = resumed_generation
        start_time = time.time()

        def callback(intermediate_result):
            nonlocal generation
            generation += 1
            elapsed = time.time() - start_time
            best_x = np.asarray(intermediate_result.x, dtype=float)
            best_power = -float(intermediate_result.fun)
            convergence = float(getattr(intermediate_result, "convergence", math.nan))
            pop = np.asarray(getattr(intermediate_result, "population", []), dtype=float)
            payload = {
                "semantic_hash": SEMANTIC_HASH,
                "data_sha256": LOCKED_DATA_SHA256,
                "compute_signature": COMPUTE_SIGNATURE,
                "stage": "de_running",
                "generation": generation,
                "best_x": best_x.tolist(),
                "best_fun": float(intermediate_result.fun),
                "best_power": best_power,
                "convergence": convergence,
                "population": pop.tolist() if pop.ndim == 2 else [],
                "nfev": int(getattr(intermediate_result, "nfev", -1)),
            }
            atomic_write_json(checkpoint_path, payload)
            progress(
                "[4/8]",
                f"DE 第 {generation}/{DE_MAXITER} 代完成 | 当前最好 a={A_MAX*best_x[0]:.6f}, "
                f"n={best_x[1]:.8f}, P≈{best_power:.6f} W | 本轮已用 {elapsed/60:.1f} min",
            )
            return False

        workers = MAX_WORKERS
        progress(
            "[4/8]",
            f"开始幂律阻尼 DE 全局搜索：workers={workers}，每代约 {DE_POPSIZE*2} 个候选，"
            f"最多 {DE_MAXITER} 代。可用环境变量 Q2_WORKERS 调整进程数。",
        )
        try:
            de = differential_evolution(
                objective, bounds=[(0.0, 1.0), (0.0, 1.0)], seed=DE_SEED,
                maxiter=remaining, popsize=DE_POPSIZE, tol=DE_TOL, atol=0.0,
                polish=False, updating="deferred", workers=workers, x0=x0,
                init=init, callback=callback,
            )
        except Exception as exc:
            if workers > 1:
                raise RuntimeError(
                    f"并行 DE 启动/执行失败（workers={workers}）。为避免静默降级，程序已停止。"
                    "如需单进程重试，请在运行前设置环境变量 Q2_WORKERS=1。"
                ) from exc
            raise
        de_payload = {
            "semantic_hash": SEMANTIC_HASH,
            "data_sha256": LOCKED_DATA_SHA256,
            "compute_signature": COMPUTE_SIGNATURE,
            "stage": "de_complete",
            "generation": resumed_generation + int(de.nit),
            "best_x": np.asarray(de.x, dtype=float).tolist(),
            "best_fun": float(de.fun),
            "best_power": -float(de.fun),
            "population": np.asarray(getattr(de, "population", []), dtype=float).tolist(),
            "nfev": int(de.nfev),
            "nit": int(de.nit),
            "success": bool(de.success),
            "message": str(de.message),
        }
        atomic_write_json(checkpoint_path, de_payload)
        progress("[4/8]", f"DE 完成：best P≈{-float(de.fun):.6f} W，checkpoint 已保存。")

    de_x = np.asarray(de_payload["best_x"], dtype=float)
    de_fun = float(de_payload["best_fun"])
    de_success = bool(de_payload.get("success", True))
    de_message = str(de_payload.get("message", "resumed from checkpoint"))
    de_nit = int(de_payload.get("nit", de_payload.get("generation", 0)))
    de_nfev = int(de_payload.get("nfev", -1))
    POWER_CACHE[(round(A_MAX * float(de_x[0]), 8), round(float(de_x[1]), 10))] = -de_fun

    local_calls = 0
    local_iter = 0

    def local_objective(x: np.ndarray) -> float:
        nonlocal local_calls
        local_calls += 1
        u = min(1.0, max(0.0, float(x[0])))
        n = min(1.0, max(0.0, float(x[1])))
        return -cached_search_power(params, A_MAX * u, n)

    def local_callback(xk: np.ndarray) -> None:
        nonlocal local_iter
        local_iter += 1
        u = float(np.clip(xk[0], 0.0, 1.0))
        n = float(np.clip(xk[1], 0.0, 1.0))
        key = (round(A_MAX * u, 8), round(n, 10))
        p = POWER_CACHE.get(key)
        suffix = f", P≈{p:.6f} W" if p is not None else ""
        progress("[5/8]", f"Powell 第 {local_iter} 次迭代：a={A_MAX*u:.6f}, n={n:.8f}{suffix}")

    progress("[5/8]", "开始 Powell 有界局部精修；重复参数会直接使用缓存。")
    local = minimize(
        local_objective, np.clip(de_x, 0.0, 1.0), method="Powell",
        bounds=[(0.0, 1.0), (0.0, 1.0)], callback=local_callback,
        options={"xtol": 1.0e-6, "ftol": 1.0e-8, "maxiter": 30, "disp": False},
    )
    if not de_success and not local.success:
        raise RuntimeError(f"DE与局部精修均未正常终止: DE={de_message}; Powell={local.message}")
    candidates = [(de_x, de_fun, "DE")]
    if np.all(np.isfinite(local.x)) and math.isfinite(float(local.fun)):
        candidates.append((np.asarray(local.x, dtype=float), float(local.fun), "DE+Powell"))
    best_x, best_fun, source = min(candidates, key=lambda item: item[1])
    best_x = np.clip(best_x, 0.0, 1.0)
    result = {
        "u": float(best_x[0]), "a": A_MAX * float(best_x[0]), "n": float(best_x[1]),
        "search_power": -best_fun, "source": source, "de_success": de_success,
        "de_message": de_message, "de_nit": de_nit, "de_nfev": de_nfev,
        "local_success": bool(local.success), "local_message": str(local.message),
        "local_nit": int(getattr(local, "nit", -1)), "local_nfev": int(getattr(local, "nfev", -1)),
        "objective_calls": max(0, de_nfev) + local_calls,
        "workers": MAX_WORKERS,
    }
    rows = [
        ["parallel differential_evolution", de_success, de_message, de_nit, de_nfev, float(de_x[0]), float(de_x[1]), -de_fun],
        ["bounded Powell refinement", local.success, local.message, getattr(local, "nit", None),
         getattr(local, "nfev", None), float(local.x[0]), float(local.x[1]), -float(local.fun)],
        ["selected", True, source, None, result["objective_calls"], result["u"], result["n"], result["search_power"]],
    ]
    progress("[5/8]", f"优化候选确定：a={result['a']:.8f}, n={result['n']:.10f}, P_search≈{result['search_power']:.6f} W")
    return result, rows


def neighborhood_check(
    params: dict[str, float], result: dict[str, object], c_linear: float,
) -> tuple[list[list[object]], float]:
    u = float(result["u"])
    n = float(result["n"])
    candidate_power = cached_search_power(params, A_MAX * u, n)
    raw_points = [
        ("u-0.01", u - 0.01, n), ("u+0.01", u + 0.01, n),
        ("n-0.01", u, n - 0.01), ("n+0.01", u, n + 0.01),
        ("diag--", u - 0.01, n - 0.01), ("diag++", u + 0.01, n + 0.01),
        ("a=0", 0.0, n), ("a=100000", 1.0, n),
        ("n=0 linear-boundary", c_linear / A_MAX, 0.0), ("n=1 projection", u, 1.0),
    ]
    rows: list[list[object]] = []
    best_alt = -math.inf
    seen: set[tuple[float, float]] = set()
    total = len(raw_points)
    for index, (label, uu, nn) in enumerate(raw_points, start=1):
        uu = min(1.0, max(0.0, float(uu)))
        nn = min(1.0, max(0.0, float(nn)))
        key = (round(uu, 12), round(nn, 12))
        if key in seen or (math.isclose(uu, u, abs_tol=1.0e-12) and math.isclose(nn, n, abs_tol=1.0e-12)):
            continue
        seen.add(key)
        progress("[7/8]", f"邻域检查 {index}/{total}: {label}, a={A_MAX*uu:.5f}, n={nn:.6f}")
        power = cached_search_power(params, A_MAX * uu, nn)
        best_alt = max(best_alt, power)
        rows.append([label, A_MAX * uu, nn, power, power - candidate_power])
    if best_alt > candidate_power * (1.0 + NEIGHBOR_TOL) + 1.0e-8:
        raise RuntimeError("边界/局部邻域发现显著高于当前非线性候选的功率，优化结果不能验收")
    return rows, candidate_power


def excel_cell_value(value: object) -> object:
    """Convert structured runtime metadata to a value openpyxl can store in one cell."""
    if isinstance(value, (list, tuple, dict, set)):
        if isinstance(value, set):
            value = sorted(value, key=str)
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, np.generic):
        return value.item()
    return value


def add_sheet(book: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append([excel_cell_value(value) for value in headers])
    for row in rows:
        sheet.append([excel_cell_value(value) for value in row])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(46, max(11, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def write_workbook(
    output: Path, code_hash: str, params: dict[str, float], audit_rows: list[list[object]],
    linear: dict[str, object], nonlinear: dict[str, object], opt_rows: list[list[object]],
    convergence_rows: list[list[object]], neighbor_rows: list[list[object]], quality_rows: list[list[object]],
) -> None:
    book = Workbook()
    book.remove(book.active)
    runtime_rows = [[key, value] for key, value in FULL_FIDELITY_CONFIG.items()]
    runtime_rows += [["code_sha256", code_hash], ["scipy_version", scipy.__version__], ["platform", platform.platform()],
                     ["fallback_used", False], ["actual_stop_reason", "all Q2 quality gates passed"]]
    add_sheet(book, "运行配置", ["项目", "值"], runtime_rows)
    add_sheet(book, "数据审计", ["状态", "检查项", "实际/结论", "期望/处理"], audit_rows)

    core_rows = [
        ["问题2入射波浪频率ω", params["omega"], "s^-1"],
        ["直线阻尼最优c", linear["c_star"], "N·s/m"],
        ["直线阻尼最大平均功率", linear["closed_power"], "W"],
        ["直线阻尼时域核验平均功率", linear["time_power"], "W"],
        ["幂律阻尼最优比例系数a", nonlinear["a"], "N·(s/m)^(n+1)"],
        ["幂律阻尼最优幂指数n", nonlinear["n"], "1"],
        ["幂律阻尼最大平均功率", nonlinear["final_power"], "W"],
        ["幂律候选最终验证停止周期", nonlinear["stop_cycle"], "T"],
    ]
    add_sheet(book, "核心指标", ["指标", "数值", "单位"], core_rows)

    coeff = linear["coeff"]
    linear_rows = [
        ["H", coeff["H"], ""], ["B", coeff["B"], ""], ["G", coeff["G"], ""],
        ["A0", coeff["A0"], ""], ["B0", coeff["B0"], ""], ["A1", coeff["A1"], ""], ["B1", coeff["B1"], ""],
        ["alpha", coeff["alpha"], ""], ["beta", coeff["beta"], ""], ["gamma", coeff["gamma"], ""],
        ["c0=sqrt(gamma/alpha)", linear["c0"], "N·s/m"], ["c*=clip(c0)", linear["c_star"], "N·s/m"],
        ["Pbar_closed(c*)", linear["closed_power"], "W"], ["Pbar_time(c*)", linear["time_power"], "W"],
        ["解析-时域相对差", linear["cross_error"], "1"],
    ]
    add_sheet(book, "直线阻尼解析", ["项目", "数值", "单位"], linear_rows)
    add_sheet(book, "优化诊断", ["阶段", "success", "message/source", "nit", "nfev/calls", "u=a/100000", "n", "search_power/W"], opt_rows)
    add_sheet(book, "收敛诊断", ["对象", "前窗口起始周期", "前窗口结束周期", "后窗口结束周期",
                                  "前窗口功率/W", "后窗口功率/W", "功率相对差", "同相位状态相对差", "通过"], convergence_rows)
    add_sheet(book, "边界邻域检查", ["检查点", "a", "n", "60T搜索口径功率/W", "相对候选差/W"], neighbor_rows)
    add_sheet(book, "主结果质量门", ["检查项", "是否通过", "证据"], quality_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)


def main() -> None:
    script_path = Path(__file__).resolve()
    project_root = script_path.parents[1]
    work_dir = project_root / "问题二求解"
    output = work_dir / "问题二求解结果.xlsx"
    de_checkpoint = work_dir / "q2_de_checkpoint.json"
    result_checkpoint = work_dir / "q2_result_checkpoint.json"
    progress("[启动]", f"Q2 full_fidelity 开始；CPU逻辑线程={os.cpu_count()}, DE workers={MAX_WORKERS}")
    progress("[启动]", "提示：可在 Windows CMD 中用 `set Q2_WORKERS=8` 调整并行进程数。")

    progress("[1/8]", "校验附件3/附件4哈希与路径。")
    _, audit_rows = audit_source_files(project_root)
    params, parameter_rows = read_q2_parameters(project_root)
    audit_rows.extend(parameter_rows)
    progress("[1/8]", f"数据通过；问题2 ω={params['omega']:.6f} s^-1。")

    cached_result = read_json_if_valid(result_checkpoint)
    if cached_result is not None and cached_result.get("stage") == "quality_gates_passed":
        progress("[缓存]", "检测到同一模型/数据/计算口径的完整计算 checkpoint，跳过昂贵优化，直接重新导出 Excel。")
        linear = cached_result["linear"]
        nonlinear = cached_result["nonlinear"]
        opt_rows = cached_result["opt_rows"]
        convergence_rows = cached_result["convergence_rows"]
        neighbor_rows = cached_result["neighbor_rows"]
        quality_rows = cached_result["quality_rows"]
        code_hash = sha256_file(script_path)
        progress("[8/8]", f"正在写入 {output.name}。")
        write_workbook(output, code_hash, params, audit_rows, linear, nonlinear, opt_rows, convergence_rows, neighbor_rows, quality_rows)
        progress("[完成]", f"已生成：{output}")
        return

    progress("[2/8]", "计算直线阻尼解析最优。")
    c_star, closed_power, coeff = linear_optimum(params)
    c0 = math.sqrt(coeff["gamma"] / coeff["alpha"])
    progress("[2/8]", f"解析结果：c*={c_star:.10f} N·s/m, P̄={closed_power:.10f} W。")

    progress("[3/8]", "用 DOP853 独立核验直线阻尼稳态平均功率。")
    linear_verify = verify_periodic_candidate(
        params, ("linear", c_star, 0.0), 40.0, require_phase=False, label="直线 c*"
    )
    linear_time_power = float(linear_verify["power"])
    linear_cross_error = abs(linear_time_power - closed_power) / (1.0 + abs(closed_power))
    if linear_cross_error > LINEAR_CROSS_TOL:
        raise RuntimeError("直线阻尼解析功率与DOP853稳态功率不一致")
    progress("[3/8]", f"交叉核验通过：P_time={linear_time_power:.10f} W，相对差={linear_cross_error:.3e}。")

    nonlinear, opt_rows = nonlinear_optimum(params, c_star, de_checkpoint)

    progress("[6/8]", "正式验证非线性最优候选（T/40）。")
    nonlinear_verify = verify_periodic_candidate(
        params, ("power", float(nonlinear["a"]), float(nonlinear["n"])), 40.0,
        require_phase=True, label="幂律候选 T/40",
    )
    progress("[6/8]", "步长加密复算非线性最优候选（T/80）。")
    refined_verify = verify_periodic_candidate(
        params, ("power", float(nonlinear["a"]), float(nonlinear["n"])), 80.0,
        require_phase=True, label="幂律候选 T/80",
    )
    step_error = abs(float(refined_verify["power"]) - float(nonlinear_verify["power"])) / (1.0 + abs(float(refined_verify["power"])))
    if step_error > STEP_REFINEMENT_TOL:
        raise RuntimeError("非线性最优候选步长加密复算未通过")
    progress("[6/8]", f"步长加密通过：相对功率差={step_error:.3e}。")

    progress("[6/8]", "检查 n=0 时幂律模型严格退化为直线阻尼。")
    n0_verify = verify_periodic_candidate(
        params, ("power", c_star, 0.0), 40.0, require_phase=False, label="n=0 等价边界"
    )
    n0_error = abs(float(n0_verify["power"]) - linear_time_power) / (1.0 + abs(linear_time_power))
    if n0_error > N0_EQUIV_TOL:
        raise RuntimeError("n=0时幂律模型未严格退化为直线阻尼模型")
    if float(nonlinear_verify["power"]) + 1.0e-6 < linear_time_power * (1.0 - NEIGHBOR_TOL):
        raise RuntimeError("非线性二维优化结果低于其合法n=0边界最优值，停止验收")
    progress("[6/8]", f"n=0 等价检查通过：相对差={n0_error:.3e}。")

    progress("[7/8]", "执行边界与局部邻域复核；重复点使用缓存。")
    neighbor_rows, candidate_search_power = neighborhood_check(params, nonlinear, c_star)
    nonlinear["final_power"] = float(nonlinear_verify["power"])
    nonlinear["refined_power"] = float(refined_verify["power"])
    nonlinear["step_error"] = step_error
    nonlinear["stop_cycle"] = int(nonlinear_verify["stop_cycle"])
    nonlinear["candidate_search_power"] = candidate_search_power

    convergence_rows: list[list[object]] = []
    for label, result in [
        ("直线c*", linear_verify), ("幂律候选", nonlinear_verify),
        ("幂律候选-步长加密", refined_verify), ("n=0等价", n0_verify),
    ]:
        for row in result["diagnostics"]:
            convergence_rows.append([label, *row])

    quality_rows = [
        ["原始附件哈希与Q2数据作用域", "是", "附件3/4 SHA-256与锁定版本一致，且读取附件3问题2行"],
        ["直线阻尼区间最优性", "是", f"c0={c0:.10f}, c*={c_star:.10f}; 已比较区间端点"],
        ["直线解析-时域交叉核验", "是", f"相对差={linear_cross_error:.3e} <= {LINEAR_CROSS_TOL:.1e}"],
        ["幂律候选参数范围", "是", f"a={nonlinear['a']:.10f}, n={nonlinear['n']:.10f} 均在题面范围内"],
        ["幂律周期稳态收敛", "是", f"最终验证至{nonlinear['stop_cycle']}T满足功率/同相位状态阈值"],
        ["幂律步长加密", "是", f"相对功率差={step_error:.3e} <= {STEP_REFINEMENT_TOL:.1e}"],
        ["n=0跨模型等价", "是", f"相对差={n0_error:.3e} <= {N0_EQUIV_TOL:.1e}"],
        ["n=0下界检查", "是", "二维幂律候选功率不低于合法线性边界最优值（数值容差内）"],
        ["边界/局部邻域检查", "是", f"{len(neighbor_rows)}个邻域/边界候选未发现超过当前候选{NEIGHBOR_TOL:.2%}的点"],
    ]
    linear = {
        "c0": c0, "c_star": c_star, "closed_power": closed_power, "time_power": linear_time_power,
        "cross_error": linear_cross_error, "coeff": coeff,
    }

    result_payload = {
        "semantic_hash": SEMANTIC_HASH,
        "data_sha256": LOCKED_DATA_SHA256,
        "compute_signature": COMPUTE_SIGNATURE,
        "stage": "quality_gates_passed",
        "linear": jsonable(linear),
        "nonlinear": jsonable(nonlinear),
        "opt_rows": jsonable(opt_rows),
        "convergence_rows": jsonable(convergence_rows),
        "neighbor_rows": jsonable(neighbor_rows),
        "quality_rows": jsonable(quality_rows),
    }
    atomic_write_json(result_checkpoint, result_payload)
    progress("[7/8]", f"全部数值质量门通过；完整结果 checkpoint 已保存：{result_checkpoint.name}")

    code_hash = sha256_file(script_path)
    progress("[8/8]", f"正在写入 Excel：{output.name}")
    write_workbook(output, code_hash, params, audit_rows, linear, nonlinear, opt_rows, convergence_rows, neighbor_rows, quality_rows)
    progress("[完成]", f"Q2 完成，已生成：{output}")


if __name__ == "__main__":
    freeze_support()
    main()
