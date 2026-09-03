#!/usr/bin/env python3
"""2022A 问题一：浮子—振子垂荡动力学主求解。

本脚本必须由用户本地 full_fidelity 执行。它读取项目 source/ 中的附件3、附件4
与题目给定 result1-1/result1-2 模板，求解两种直线阻尼情形，并输出：
1) 问题一求解/result1-1.xlsx
2) 问题一求解/result1-2.xlsx
3) 问题一求解/问题一求解结果.xlsx（运行配置、审计与质量门证据）

模型语义绑定：Q1 semantic revision=1,
hash=3c949a24b44d6d9ee12abe4e763e47942406456c523faf295263f06fac297192。
"""
from __future__ import annotations

import hashlib
import logging
import math
import platform
from pathlib import Path

import numpy as np
import scipy
from openpyxl import Workbook, load_workbook
from scipy.integrate import solve_ivp
from scipy.linalg import expm


FULL_FIDELITY_CONFIG = {
    "execution_owner": "user",
    "execution_profile": "full_fidelity",
    "stage": "primary",
    "problem_name": "问题一",
    "data_paths": [
        "source/附件3.xlsx",
        "source/附件4.xlsx",
        "source/result1-1.xlsx",
        "source/result1-2.xlsx",
    ],
    "data_sha256": "6b92eac92ef46cb507b97f11a9a555bc60da394c4efc2e6281df631e52bd33e6",
    "solver": "scipy.integrate.solve_ivp(method='DOP853')",
    "solver_version": "SciPy runtime version is recorded in workbook",
    "random_seed": "not_applicable_deterministic_ode",
    "tolerance": 1.0e-10,
    "iteration_or_time_limit": "integrate from 0 to 40T; no reduced horizon or wall-time cutoff",
    "expected_workbook": "问题一求解/问题一求解结果.xlsx",
    "allow_reduced_data": False,
    "allow_coarser_grid": False,
    "allow_shorter_horizon": False,
    "allow_fewer_repetitions": False,
    "allow_relaxed_tolerance": False,
    "allow_silent_solver_fallback": False,
}

EXPECTED_HASHES = {
    "附件3.xlsx": "50a5dd70f04dfb0a57fb2602422dc7999b30aad54ddc02353f5b8f01423fd612",
    "附件4.xlsx": "c8eff812f5980d955b4f0e587c5f7a357b2571d8d903fcb4913fba77c7354d6d",
    "result1-1.xlsx": "83ed6e0f2ebcdbdcb53e99a3bfebfbd8dc16141f91396eba8806e781d7809c7a",
    "result1-2.xlsx": "cc0abbceff32f425e738a3d9c0534fc3fbab4b2a1d2d86b8dc4d51229fb820bf",
}

RTOL = 1.0e-10
ATOL = 1.0e-12
OUTPUT_DT = 0.2
CHECK_TIMES = (10.0, 20.0, 40.0, 60.0, 100.0)
LINEAR_VALIDATION_TOL = 1.0e-8
NONLINEAR_VALIDATION_TOL = 1.0e-8


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def aggregate_data_hash(actual_hashes: dict[str, str]) -> str:
    text = "".join(f"{name}:{actual_hashes[name]}\n" for name in sorted(actual_hashes))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def audit_source_files(source_dir: Path) -> tuple[dict[str, str], list[list[object]]]:
    actual: dict[str, str] = {}
    rows: list[list[object]] = []
    for name, expected in EXPECTED_HASHES.items():
        path = source_dir / name
        if not path.is_file():
            raise FileNotFoundError(f"缺少输入文件: {path}")
        current = sha256_file(path)
        actual[name] = current
        ok = current.lower() == expected.lower()
        rows.append(["阻断" if not ok else "通过", f"{name} SHA-256", current, "保持原始文件，不做变换"])
        if not ok:
            raise ValueError(f"{name} 哈希与锁定附件不一致，停止求解")
    aggregate = aggregate_data_hash(actual)
    if aggregate != FULL_FIDELITY_CONFIG["data_sha256"]:
        raise ValueError("聚合 data_sha256 与 FULL_FIDELITY_CONFIG 不一致")
    return actual, rows


def read_parameter_tables(source_dir: Path) -> tuple[dict[str, float], list[list[object]]]:
    wb3 = load_workbook(source_dir / "附件3.xlsx", data_only=True, read_only=True)
    ws3 = wb3.active
    headers3 = [ws3.cell(1, col).value for col in range(1, 9)]
    expected3 = [
        "问题", "入射波浪频率 (s-1)", "垂荡附加质量 (kg)", "纵摇附加转动惯量 (kg·m2)",
        "垂荡兴波阻尼系数 (N·s/m)", "纵摇兴波阻尼系数 (N·m·s)",
        "垂荡激励力振幅 (N)", "纵摇激励力矩振幅 (N·m)",
    ]
    if headers3 != expected3:
        raise ValueError("附件3表头与锁定版本不一致")
    q1_row = next((row for row in ws3.iter_rows(min_row=2, values_only=True) if row[0] == "问题1"), None)
    if q1_row is None:
        raise ValueError("附件3中未找到问题1参数行")

    wb4 = load_workbook(source_dir / "附件4.xlsx", data_only=True, read_only=True)
    ws4 = wb4.active
    if [ws4["A1"].value, ws4["B1"].value] != ["参数", "取值"]:
        raise ValueError("附件4表头与锁定版本不一致")
    p4 = {str(row[0]): float(row[1]) for row in ws4.iter_rows(min_row=2, values_only=True) if row[0] is not None}
    needed = ["浮子质量 (kg)", "浮子底半径 (m)", "振子质量 (kg)", "海水的密度 (kg/m3)", "重力加速度 (m/s2)", "弹簧刚度 (N/m)"]
    if any(name not in p4 for name in needed):
        raise ValueError("附件4缺少Q1所需字段")

    params = {
        "omega": float(q1_row[1]), "added_mass": float(q1_row[2]),
        "wave_damping": float(q1_row[4]), "force_amp": float(q1_row[6]),
        "float_mass": p4["浮子质量 (kg)"], "radius": p4["浮子底半径 (m)"],
        "osc_mass": p4["振子质量 (kg)"], "rho": p4["海水的密度 (kg/m3)"],
        "g": p4["重力加速度 (m/s2)"], "spring_k": p4["弹簧刚度 (N/m)"],
    }
    if not all(math.isfinite(value) for value in params.values()):
        raise ValueError("Q1参数存在NaN或Inf")
    rows = [
        ["通过", "附件3表头", "8个字段与锁定版本一致", "不修改"],
        ["通过", "附件3问题1参数", "频率/附加质量/兴波阻尼/激励力均为有限数", "直接用于模型"],
        ["通过", "附件4字段", "Q1所需6个物理/几何字段齐全且为有限数", "直接用于模型"],
    ]
    return params, rows


def validate_templates(source_dir: Path) -> list[list[object]]:
    expected = [["时间 (s)", "浮子", None, "振子", None], [None, "位移 (m)", "速度 (m/s)", "位移 (m)", "速度 (m/s)"]]
    rows: list[list[object]] = []
    for name in ("result1-1.xlsx", "result1-2.xlsx"):
        wb = load_workbook(source_dir / name, data_only=False, read_only=True)
        ws = wb.active
        current = [[ws.cell(r, c).value for c in range(1, 6)] for r in (1, 2)]
        if current != expected:
            raise ValueError(f"{name}模板表头与题目附件不一致")
        rows.append(["通过", f"{name}模板", "两层表头与题目附件一致", "保留模板结构后填数"])
    return rows


def time_grid(omega: float) -> tuple[float, float, np.ndarray]:
    period = 2.0 * math.pi / omega
    t_end = 40.0 * period
    last = math.floor((t_end + 1.0e-12) / OUTPUT_DT) * OUTPUT_DT
    count = int(round(last / OUTPUT_DT)) + 1
    return period, t_end, np.linspace(0.0, last, count)


def rhs_factory(params: dict[str, float], case: str):
    mass = params["float_mass"] + params["added_mass"]
    osc_mass = params["osc_mass"]
    spring_k = params["spring_k"]
    hydro_k = params["rho"] * params["g"] * math.pi * params["radius"] ** 2
    damping = params["wave_damping"]
    force_amp = params["force_amp"]
    omega = params["omega"]

    def rhs(t: float, y: np.ndarray) -> np.ndarray:
        zf, vf, zo, vo = y
        vr = vf - vo
        fd = 10000.0 * vr if case == "constant" else 10000.0 * abs(vr) ** 0.5 * vr
        af = (force_amp * math.cos(omega * t) - damping * vf - hydro_k * zf - spring_k * (zf - zo) - fd) / mass
        ao = (spring_k * (zf - zo) + fd) / osc_mass
        return np.array([vf, af, vo, ao], dtype=float)

    return rhs


def solve_case(params: dict[str, float], t_eval: np.ndarray, t_end: float, case: str, max_step: float):
    sol = solve_ivp(
        rhs_factory(params, case), (0.0, t_end), np.zeros(4), method="DOP853",
        t_eval=t_eval, rtol=RTOL, atol=ATOL, max_step=max_step,
    )
    if not sol.success or sol.y.shape != (4, len(t_eval)):
        raise RuntimeError(f"{case}求解失败: {sol.message}")
    if not np.all(np.isfinite(sol.y)):
        raise RuntimeError(f"{case}求解结果含NaN或Inf")
    return sol


def linear_reference(params: dict[str, float], t_eval: np.ndarray) -> np.ndarray:
    mass = params["float_mass"] + params["added_mass"]
    mo = params["osc_mass"]
    k = params["spring_k"]
    c = 10000.0
    kh = params["rho"] * params["g"] * math.pi * params["radius"] ** 2
    b = params["wave_damping"]
    f = params["force_amp"]
    omega = params["omega"]
    matrix = np.zeros((6, 6), dtype=float)
    matrix[0, 1] = 1.0
    matrix[1, :4] = [-(kh + k) / mass, -(b + c) / mass, k / mass, c / mass]
    matrix[1, 4] = f / mass
    matrix[2, 3] = 1.0
    matrix[3, :4] = [k / mo, c / mo, -k / mo, -c / mo]
    matrix[4, 5] = -omega
    matrix[5, 4] = omega
    phi = expm(matrix * OUTPUT_DT)
    state = np.array([0.0, 0.0, 0.0, 0.0, 1.0, 0.0])
    out = np.empty((4, len(t_eval)), dtype=float)
    out[:, 0] = state[:4]
    for index in range(1, len(t_eval)):
        state = phi @ state
        out[:, index] = state[:4]
    return out


def normalized_error(reference: np.ndarray, candidate: np.ndarray) -> tuple[float, float]:
    absolute = float(np.max(np.abs(reference - candidate)))
    scale = 1.0 + float(np.max(np.abs(reference)))
    return absolute, absolute / scale


def designated_rows(t_eval: np.ndarray, case_data: dict[str, np.ndarray]) -> list[list[object]]:
    rows: list[list[object]] = []
    for case_name, values in case_data.items():
        for target in CHECK_TIMES:
            index = int(round(target / OUTPUT_DT))
            if index >= len(t_eval) or not math.isclose(float(t_eval[index]), target, abs_tol=1.0e-12):
                raise RuntimeError(f"指定时刻 {target}s 不在0.2s输出网格中")
            rows.append([case_name, target, *[float(values[j, index]) for j in range(4)]])
    return rows


def write_official_template(template: Path, output: Path, t_eval: np.ndarray, values: np.ndarray) -> None:
    wb = load_workbook(template)
    ws = wb.active
    for row_index, time_value in enumerate(t_eval, start=3):
        row = [float(time_value), *[float(values[j, row_index - 3]) for j in range(4)]]
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row_index, col_index, value)
            cell.number_format = "0.000000"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def add_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for column in ws.columns:
        width = min(40, max(10, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2))
        ws.column_dimensions[column[0].column_letter].width = width


def write_evidence_workbook(output: Path, code_hash: str, context: dict[str, object]) -> bool:
    params = context["params"]
    period = context["period"]
    t_end = context["t_end"]
    t_eval = context["t_eval"]
    constant = context["constant"]
    nonlinear = context["nonlinear"]
    audit_rows = context["audit_rows"]
    linear_err = context["linear_err"]
    nonlinear_err = context["nonlinear_err"]
    stop_reason = context["stop_reason"]
    wb = Workbook()
    wb.remove(wb.active)
    runtime_rows = [["execution_owner", "user"], ["execution_profile", "full_fidelity"], ["stage", "primary"],
        ["problem_name", "问题一"], ["code_sha256", code_hash], ["data_sha256", FULL_FIDELITY_CONFIG["data_sha256"]],
        ["solver", FULL_FIDELITY_CONFIG["solver"]], ["solver_version", scipy.__version__], ["tolerance", RTOL],
        ["iteration_or_time_limit", FULL_FIDELITY_CONFIG["iteration_or_time_limit"]], ["actual_stop_reason", stop_reason],
        ["random_seed", "not_applicable_deterministic_ode"], ["repetitions_or_scenarios", 2],
        ["grid_or_time_range", f"0:0.2:{t_eval[-1]:.1f}s; integration end=40T={t_end:.12f}s"], ["fallback_used", False],
        ["platform", platform.platform()], ["allow_reduced_data", False], ["allow_coarser_grid", False],
        ["allow_shorter_horizon", False], ["allow_fewer_repetitions", False], ["allow_relaxed_tolerance", False],
        ["allow_silent_solver_fallback", False]]
    add_sheet(wb, "运行配置", ["项目", "值"], runtime_rows)

    hydro_k = params["rho"] * params["g"] * math.pi * params["radius"] ** 2
    core_rows = [["波浪周期T", period, "s", "2π/ω", ""], ["40个周期终点", t_end, "s", "40T", ""],
        ["最后输出时刻", float(t_eval[-1]), "s", "0.2s等间距且不超过40T", ""], ["输出点数", len(t_eval), "个", "含t=0", ""],
        ["静水恢复刚度", hydro_k, "N/m", "ρgπR²", ""], ["Case1独立核验最大绝对误差", linear_err[0], "状态量单位", "矩阵指数 vs DOP853", ""],
        ["Case1独立核验归一化误差", linear_err[1], "1", "max|Δ|/(1+max|ref|)", ""],
        ["Case2步长加密归一化误差", nonlinear_err[1], "1", "T/200 vs T/400", ""]]
    add_sheet(wb, "核心指标", ["指标", "数值", "单位", "统计口径", "说明"], core_rows)
    add_sheet(wb, "数据审计", ["等级", "检查项", "信息", "处理方式"], audit_rows)

    dt_error = float(np.max(np.abs(np.diff(t_eval) - OUTPUT_DT)))
    constraint_rows = [["C1", "初始状态为零", 0.0, 1.0e-12, "是"], ["C2", "输出时间间隔0.2s", dt_error, 1.0e-12, "是" if dt_error <= 1.0e-12 else "否"],
        ["C3", "最后输出时刻不超过40T", max(0.0, float(t_eval[-1] - t_end)), 1.0e-12, "是"],
        ["C4", "全部状态为有限数", 0.0 if np.all(np.isfinite(constant)) and np.all(np.isfinite(nonlinear)) else 1.0, 0.0, "是" if np.all(np.isfinite(constant)) and np.all(np.isfinite(nonlinear)) else "否"]]
    add_sheet(wb, "约束违反检查", ["约束编号", "约束含义", "违反量", "容差", "是否满足"], constraint_rows)
    add_sheet(wb, "离散精度", ["离散参数", "取值", "目标指标", "相对变化"], [["输出时间间隔", OUTPUT_DT, "等间距误差", dt_error], ["主积分max_step", period / 200.0, "Case2加密比较归一化误差", nonlinear_err[1]]])
    add_sheet(wb, "收敛诊断", ["迭代或样本数", "指标", "数值", "判定"], [["Case1: DOP853 vs 矩阵指数", "归一化误差", linear_err[1], "通过" if linear_err[1] <= LINEAR_VALIDATION_TOL else "未通过"], ["Case2: max_step T/200 vs T/400", "归一化误差", nonlinear_err[1], "通过" if nonlinear_err[1] <= NONLINEAR_VALIDATION_TOL else "未通过"]])

    detail_rows: list[list[object]] = []
    labels = ("浮子位移", "浮子速度", "振子位移", "振子速度")
    units = ("m", "m/s", "m", "m/s")
    for scenario, values in (("常阻尼c=10000", constant), ("幂律阻尼a=10000,n=0.5", nonlinear)):
        for i, time_value in enumerate(t_eval):
            for j, label in enumerate(labels):
                detail_rows.append([f"{scenario}-{i:04d}-{j+1}", scenario, float(time_value), float(values[j, i]), label, units[j]])
    add_sheet(wb, "仿真明细", ["记录键", "场景", "时刻", "数值", "状态量", "单位"], detail_rows)
    add_sheet(wb, "指定时刻结果", ["场景", "时间(s)", "浮子位移(m)", "浮子速度(m/s)", "振子位移(m)", "振子速度(m/s)"], designated_rows(t_eval, {"常阻尼": constant, "幂律阻尼": nonlinear}))

    quality_rows = [["输入附件哈希与锁定版本一致", "是", FULL_FIDELITY_CONFIG["data_sha256"]],
        ["两类ODE求解均成功且无fallback", "是", stop_reason], ["Case1独立矩阵指数核验", "是" if linear_err[1] <= LINEAR_VALIDATION_TOL else "否", f"归一化误差={linear_err[1]:.3e}"],
        ["Case2步长加密收敛核验", "是" if nonlinear_err[1] <= NONLINEAR_VALIDATION_TOL else "否", f"归一化误差={nonlinear_err[1]:.3e}"],
        ["0.2s输出网格约束", "是" if dt_error <= 1.0e-12 else "否", f"最大间隔误差={dt_error:.3e}"], ["结果无NaN/Inf", "是", "两场景全部有限"]]
    add_sheet(wb, "主结果质量门", ["检查项", "是否通过", "证据"], quality_rows)
    passed = all(row[1] == "是" for row in quality_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return passed


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    source_dir = project_root / "source"
    output_dir = project_root / "问题一求解"
    code_hash = sha256_file(Path(__file__).resolve())
    _, audit_rows = audit_source_files(source_dir)
    params, parameter_audit = read_parameter_tables(source_dir)
    audit_rows.extend(parameter_audit)
    audit_rows.extend(validate_templates(source_dir))
    period, t_end, t_eval = time_grid(params["omega"])
    max_step = period / 200.0

    constant_sol = solve_case(params, t_eval, t_end, "constant", max_step)
    nonlinear_sol = solve_case(params, t_eval, t_end, "nonlinear", max_step)
    linear_ref = linear_reference(params, t_eval)
    nonlinear_refined = solve_case(params, t_eval, t_end, "nonlinear", period / 400.0)
    linear_err = normalized_error(linear_ref, constant_sol.y)
    nonlinear_err = normalized_error(nonlinear_refined.y, nonlinear_sol.y)
    stop_reason = f"constant={constant_sol.message}; nonlinear={nonlinear_sol.message}; fallback=false"

    write_official_template(source_dir / "result1-1.xlsx", output_dir / "result1-1.xlsx", t_eval, constant_sol.y)
    write_official_template(source_dir / "result1-2.xlsx", output_dir / "result1-2.xlsx", t_eval, nonlinear_sol.y)
    context = {
        "params": params, "period": period, "t_end": t_end, "t_eval": t_eval,
        "constant": constant_sol.y, "nonlinear": nonlinear_sol.y, "audit_rows": audit_rows,
        "linear_err": linear_err, "nonlinear_err": nonlinear_err, "stop_reason": stop_reason,
    }
    quality_passed = write_evidence_workbook(output_dir / "问题一求解结果.xlsx", code_hash, context)
    if not quality_passed:
        raise RuntimeError("主结果质量门未通过；请返回问题一求解结果.xlsx检查失败证据，不得进入下游")
    logging.info("Q1 full_fidelity完成：已生成两份题目模板结果与标准主结果工作簿。")


if __name__ == "__main__":
    main()
