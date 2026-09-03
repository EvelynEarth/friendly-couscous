#!/usr/bin/env python3
"""2022A 问题一：主结果通过质量门后的独立深化分析。

本脚本不重新求解 Q1 ODE，只读取已验收的“问题一求解结果.xlsx”，
围绕三个评委风险开展题目专属分析：
1. 40 个波浪周期末段是否已经形成稳定周期响应；
2. 常阻尼与幂律阻尼两种题设结构是否保持同一“有界周期响应”定性结论；
3. 两种阻尼的局部差异出现在哪里，Figure Evidence 的 Local Zoom 应放大哪个真实区间。

脚本由用户本地 full_fidelity 执行，输出：
    问题一求解/问题一结果深化分析.xlsx
"""
from __future__ import annotations

import hashlib
import math
import platform
from copy import copy
from pathlib import Path

import numpy as np
import openpyxl


LOCKED_DATA_SHA256 = "6b92eac92ef46cb507b97f11a9a555bc60da394c4efc2e6281df631e52bd33e6"
PRIMARY_WORKBOOK_SHA256 = "c9d2e299c63269188a15ade091270ff4fdf0f695a2235d1669e96cece288448b"
TAIL_CYCLES = (35, 36, 37, 38, 39)
STABILITY_TOL = 0.02
ROI_WINDOW_SECONDS = 2.4
EPS = 1.0e-12

FULL_FIDELITY_CONFIG = {
    "execution_owner": "user",
    "execution_profile": "full_fidelity",
    "stage": "analysis",
    "problem_name": "问题一",
    "data_paths": ["问题一求解/问题一求解结果.xlsx"],
    "data_sha256": LOCKED_DATA_SHA256,
    "solver": "deterministic post-processing of accepted primary workbook",
    "solver_version": "NumPy runtime version is recorded in workbook",
    "random_seed": "not_applicable_deterministic_analysis",
    "tolerance": STABILITY_TOL,
    "iteration_or_time_limit": "use all 898 accepted samples and all 39 complete wave cycles; no reduced horizon",
    "expected_workbook": "问题一求解/问题一结果深化分析.xlsx",
    "allow_reduced_data": False,
    "allow_coarser_grid": False,
    "allow_shorter_horizon": False,
    "allow_fewer_repetitions": False,
    "allow_relaxed_tolerance": False,
    "allow_silent_solver_fallback": False,
}

SCENARIOS = ("常阻尼c=10000", "幂律阻尼a=10000,n=0.5")
STATE_LABELS = ("浮子位移", "浮子速度", "振子位移", "振子速度")
STATE_UNITS = {"浮子位移": "m", "浮子速度": "m/s", "振子位移": "m", "振子速度": "m/s"}
METRIC_LABELS = (
    "浮子位移半峰峰值",
    "振子位移半峰峰值",
    "相对位移半峰峰值",
    "相对速度RMS",
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_primary_workbook(path: Path) -> openpyxl.Workbook:
    if not path.is_file():
        raise FileNotFoundError(f"缺少已验收主工作簿: {path}")
    actual_hash = sha256_file(path)
    if actual_hash != PRIMARY_WORKBOOK_SHA256:
        raise ValueError("问题一求解结果.xlsx 与已验收版本哈希不一致，停止深化分析")
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    required = {
        "运行配置", "核心指标", "数据审计", "约束违反检查", "离散精度",
        "收敛诊断", "仿真明细", "指定时刻结果", "主结果质量门",
    }
    missing = sorted(required - set(book.sheetnames))
    if missing:
        raise ValueError(f"主工作簿缺少工作表: {missing}")
    return book


def mapping_from_two_column_sheet(book: openpyxl.Workbook, sheet: str) -> dict[str, object]:
    rows = list(book[sheet].iter_rows(values_only=True))
    if not rows or tuple(rows[0][:2]) != ("项目", "值"):
        raise ValueError(f"{sheet}表头不是项目|值")
    return {
        str(row[0]).strip(): row[1]
        for row in rows[1:]
        if row and row[0] not in (None, "")
    }


def validate_primary_evidence(book: openpyxl.Workbook) -> tuple[float, int]:
    runtime = mapping_from_two_column_sheet(book, "运行配置")
    if runtime.get("execution_owner") != "user" or runtime.get("execution_profile") != "full_fidelity":
        raise ValueError("主工作簿不是用户 full_fidelity 执行证据")
    if runtime.get("stage") != "primary" or runtime.get("problem_name") != "问题一":
        raise ValueError("主工作簿阶段或问题编号不一致")
    if str(runtime.get("data_sha256")) != LOCKED_DATA_SHA256:
        raise ValueError("主工作簿 data_sha256 与 Q1 锁定数据不一致")
    if runtime.get("fallback_used") is not False:
        raise ValueError("主工作簿记录了 solver fallback")

    gate_rows = list(book["主结果质量门"].iter_rows(values_only=True))
    if tuple(gate_rows[0][:3]) != ("检查项", "是否通过", "证据"):
        raise ValueError("主结果质量门表头不一致")
    failed = [row[0] for row in gate_rows[1:] if row and str(row[1]).strip() != "是"]
    if failed:
        raise ValueError(f"主结果质量门存在未通过项: {failed}")

    core_rows = list(book["核心指标"].iter_rows(values_only=True))
    headers = [str(item).strip() if item is not None else "" for item in core_rows[0]]
    metric_idx = headers.index("指标")
    value_idx = headers.index("数值")
    metrics = {
        str(row[metric_idx]): row[value_idx]
        for row in core_rows[1:]
        if row and row[metric_idx] not in (None, "")
    }
    period = float(metrics["波浪周期T"])
    output_points = int(metrics["输出点数"])
    if not math.isfinite(period) or period <= 0 or output_points != 898:
        raise ValueError("主工作簿周期或输出点数异常")
    return period, output_points


def read_simulation_detail(book: openpyxl.Workbook) -> dict[str, dict[str, tuple[np.ndarray, np.ndarray]]]:
    rows = list(book["仿真明细"].iter_rows(values_only=True))
    expected = ["记录键", "场景", "时刻", "数值", "状态量", "单位"]
    headers = [str(item).strip() if item is not None else "" for item in rows[0]]
    if headers != expected:
        raise ValueError("仿真明细表头与 Q1 主工作簿约定不一致")

    buckets: dict[str, dict[str, list[tuple[float, float]]]] = {
        scenario: {label: [] for label in STATE_LABELS} for scenario in SCENARIOS
    }
    for row in rows[1:]:
        if not row or row[1] not in buckets or row[4] not in STATE_LABELS:
            continue
        time_value = float(row[2])
        state_value = float(row[3])
        if not math.isfinite(time_value) or not math.isfinite(state_value):
            raise ValueError("仿真明细包含 NaN/Inf")
        buckets[str(row[1])][str(row[4])].append((time_value, state_value))

    output: dict[str, dict[str, tuple[np.ndarray, np.ndarray]]] = {}
    for scenario in SCENARIOS:
        output[scenario] = {}
        reference_times: np.ndarray | None = None
        for label in STATE_LABELS:
            pairs = sorted(buckets[scenario][label])
            if len(pairs) != 898:
                raise ValueError(f"{scenario}-{label}记录数不是898")
            times = np.array([pair[0] for pair in pairs], dtype=float)
            values = np.array([pair[1] for pair in pairs], dtype=float)
            if reference_times is None:
                reference_times = times
            elif not np.array_equal(reference_times, times):
                raise ValueError(f"{scenario}四个状态量时间网格不一致")
            output[scenario][label] = (times, values)
    return output


def half_peak_to_peak(values: np.ndarray) -> float:
    return 0.5 * float(np.max(values) - np.min(values))


def cycle_metrics(
    data: dict[str, dict[str, tuple[np.ndarray, np.ndarray]]], period: float
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for scenario in SCENARIOS:
        times = data[scenario]["浮子位移"][0]
        zf = data[scenario]["浮子位移"][1]
        vf = data[scenario]["浮子速度"][1]
        zo = data[scenario]["振子位移"][1]
        vo = data[scenario]["振子速度"][1]
        for cycle in range(1, 40):
            start = (cycle - 1) * period
            end = cycle * period
            mask = (times >= start) & (times < end)
            if int(np.count_nonzero(mask)) < 10:
                raise ValueError(f"第{cycle}周期有效采样点过少")
            rel_z = zf[mask] - zo[mask]
            rel_v = vf[mask] - vo[mask]
            metrics = {
                "浮子位移半峰峰值": half_peak_to_peak(zf[mask]),
                "振子位移半峰峰值": half_peak_to_peak(zo[mask]),
                "相对位移半峰峰值": half_peak_to_peak(rel_z),
                "相对速度RMS": float(np.sqrt(np.mean(rel_v ** 2))),
            }
            records.append({"场景": scenario, "周期": cycle, **metrics})
    return records


def tail_stability(records: list[dict[str, object]]) -> tuple[list[list[object]], dict[str, dict[str, float]], dict[str, bool]]:
    detail_rows: list[list[object]] = []
    references: dict[str, dict[str, float]] = {}
    passed: dict[str, bool] = {}
    for scenario in SCENARIOS:
        selected = [row for row in records if row["场景"] == scenario and row["周期"] in TAIL_CYCLES]
        references[scenario] = {}
        metric_deviations: list[float] = []
        for metric in METRIC_LABELS:
            values = np.array([float(row[metric]) for row in selected], dtype=float)
            reference = float(np.mean(values))
            max_dev = float(np.max(np.abs(values - reference) / (abs(reference) + EPS)))
            references[scenario][metric] = reference
            metric_deviations.append(max_dev)
            detail_rows.append([scenario, metric, reference, max_dev, STABILITY_TOL, "通过" if max_dev <= STABILITY_TOL else "未通过"])
        passed[scenario] = max(metric_deviations) <= STABILITY_TOL
    return detail_rows, references, passed


def structure_comparison(
    references: dict[str, dict[str, float]], passed: dict[str, bool]
) -> tuple[list[list[object]], list[list[object]], bool]:
    constant = SCENARIOS[0]
    nonlinear = SCENARIOS[1]
    rows: list[list[object]] = []
    relative_differences: list[list[object]] = []
    for metric in METRIC_LABELS:
        base = references[constant][metric]
        alt = references[nonlinear][metric]
        rel_diff = abs(alt - base) / (abs(base) + EPS)
        relative_differences.append([metric, base, alt, rel_diff])
    both_stable = passed[constant] and passed[nonlinear]
    rows.append([
        "常阻尼c=10000", "题设常量阻尼结构；第35-39完整周期",
        0.0, 0.0, "是" if passed[constant] else "否", "作为结构对照基准",
    ])
    rows.append([
        "幂律阻尼a=10000,n=0.5", "题设幂律阻尼结构；第35-39完整周期",
        max(item[3] for item in relative_differences),
        max(item[3] for item in relative_differences),
        "是" if both_stable else "否",
        "不要求与常阻尼数值接近；只检验两种题设结构是否都保持稳定周期尾段",
    ])
    return rows, relative_differences, both_stable


def difference_scan(
    data: dict[str, dict[str, tuple[np.ndarray, np.ndarray]]]
) -> tuple[list[list[object]], list[list[object]], tuple[float, float, float]]:
    times = data[SCENARIOS[0]][STATE_LABELS[0]][0]
    normalized_differences: list[np.ndarray] = []
    peak_rows: list[list[object]] = []
    detail_columns: list[np.ndarray] = [times]

    for label in STATE_LABELS:
        constant = data[SCENARIOS[0]][label][1]
        nonlinear = data[SCENARIOS[1]][label][1]
        diff = nonlinear - constant
        abs_diff = np.abs(diff)
        index = int(np.argmax(abs_diff))
        peak = float(abs_diff[index])
        rms = float(np.sqrt(np.mean(diff ** 2)))
        span = float(max(np.ptp(constant), np.ptp(nonlinear), EPS))
        peak_rows.append([
            label, STATE_UNITS[label], float(times[index]), peak, float(diff[index]),
            float(constant[index]), float(nonlinear[index]), rms, peak / span,
        ])
        normalized_differences.append(abs_diff / max(peak, EPS))
        detail_columns.extend([constant, nonlinear, diff])

    score = np.mean(np.vstack(normalized_differences), axis=0)
    detail_columns.append(score)
    dt = float(times[1] - times[0])
    window_points = int(round(ROI_WINDOW_SECONDS / dt)) + 1
    rolling = np.convolve(score, np.ones(window_points) / window_points, mode="valid")
    start_index = int(np.argmax(rolling))
    end_index = start_index + window_points - 1
    roi = (float(times[start_index]), float(times[end_index]), float(rolling[start_index]))
    detail_rows = np.column_stack(detail_columns).tolist()
    return peak_rows, detail_rows, roi


def add_sheet(book: openpyxl.Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = book.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(42, max(11, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def write_analysis_workbook(output: Path, code_hash: str, context: dict[str, object]) -> None:
    source_hash = str(context["source_hash"])
    records = context["records"]
    stability_rows = context["stability_rows"]
    stable = context["stable"]
    structure_rows = context["structure_rows"]
    differences = context["differences"]
    both_stable = bool(context["both_stable"])
    peak_rows = context["peak_rows"]
    diff_detail = context["diff_detail"]
    roi_start, roi_end, roi_score = context["roi"]
    book = openpyxl.Workbook()
    book.remove(book.active)

    runtime_rows = [
        ["execution_owner", "user"], ["execution_profile", "full_fidelity"], ["stage", "analysis"],
        ["problem_name", "问题一"], ["code_sha256", code_hash], ["data_sha256", LOCKED_DATA_SHA256],
        ["source_workbook_sha256", source_hash], ["solver", FULL_FIDELITY_CONFIG["solver"]],
        ["solver_version", np.__version__], ["tolerance", STABILITY_TOL],
        ["iteration_or_time_limit", FULL_FIDELITY_CONFIG["iteration_or_time_limit"]],
        ["actual_stop_reason", "analysis completed; fallback=false"],
        ["random_seed", "not_applicable_deterministic_analysis"], ["repetitions_or_scenarios", 2],
        ["grid_or_time_range", "accepted primary workbook: 0:0.2:179.4s; analyze all 898 points and complete cycles 1-39"],
        ["fallback_used", False], ["platform", platform.platform()], ["allow_reduced_data", False],
        ["allow_coarser_grid", False], ["allow_shorter_horizon", False], ["allow_fewer_repetitions", False],
        ["allow_relaxed_tolerance", False], ["allow_silent_solver_fallback", False],
    ]
    add_sheet(book, "运行配置", ["项目", "值"], runtime_rows)

    design_rows = [
        ["40周期结果可能仍混有瞬态", "末段是否已形成稳定周期响应",
         "按波浪周期分段，比较第35-39个完整周期的关键幅值与相对速度RMS",
         "五周期内各指标相对其均值的最大相对偏差", "≤2%",
         "问题一求解结果.xlsx/仿真明细", "证明40周期时域足以形成稳定尾段", "先验固定2%周期重复性阈值"],
        ["阻尼律结构变化可能改变定性结论", "两种题设阻尼律是否都保持有界稳定周期响应",
         "分别检验两种阻尼律末5个完整周期稳定性，并报告稳态指标相对差异",
         "两场景末段稳定判定", "两场景均通过", "问题一求解结果.xlsx/仿真明细",
         "支持两种题设情形的对比表述", "不要求两种阻尼数值接近"],
        ["全时域曲线高度相似使局部差异被压缩", "Local Zoom 应放大哪个真实时间窗",
         "逐状态计算幂律-常阻尼差值；按各自最大差归一化后求综合差异分数；在2.4s滑动窗上取均值最大区间",
         "四状态峰值时刻、最大绝对差、综合滑窗分数", "ROI完全由真实时序决定且覆盖差异高信息区",
         "问题一求解结果.xlsx/仿真明细", "为q1_plot.m提供可复核的差异驱动局部放大区", "避免任意截取或放大无差异区域"],
    ]
    add_sheet(book, "分析设计", ["风险来源", "分析问题", "方法", "指标", "通过标准", "输入", "论文作用", "选择理由"], design_rows)

    cycle_rows = [[row["场景"], row["周期"], row["浮子位移半峰峰值"], row["振子位移半峰峰值"],
                   row["相对位移半峰峰值"], row["相对速度RMS"]] for row in records]
    add_sheet(book, "周期指标明细", ["场景", "周期", *METRIC_LABELS], cycle_rows)
    add_sheet(book, "误差分解", ["误差来源", "指标", "数值", "占比", "分组", "时间范围", "说明"], [
        ["末5周期重复性", row[1], row[3], row[3], row[0], "第35-39完整周期", f"阈值={STABILITY_TOL:.0%}; {row[5]}"]
        for row in stability_rows
    ])
    add_sheet(book, "结构稳健性", ["替代结构", "核心设定", "结果指标", "与主模型差异", "结论是否一致", "说明"], structure_rows)
    add_sheet(book, "稳态结构差异", ["指标", "常阻尼稳态值", "幂律阻尼稳态值", "相对差异"], differences)
    add_sheet(book, "局部差异峰值", ["状态量", "单位", "峰值时刻(s)", "最大绝对差", "有符号差", "常阻尼值", "幂律阻尼值", "全时域RMS差", "峰值/全局跨度"], peak_rows)
    add_sheet(book, "局部放大ROI", ["对象", "起点(s)", "终点(s)", "宽度(s)", "综合差异分数", "依据"], [[
        "共同Local Zoom ROI", roi_start, roi_end, roi_end - roi_start, roi_score,
        "四状态绝对差按各自峰值归一化后取平均，2.4s滑动窗均值最大",
    ]])
    add_sheet(book, "差异时序", [
        "时间(s)", "常阻尼浮子位移(m)", "幂律阻尼浮子位移(m)", "浮子位移差(m)",
        "常阻尼浮子速度(m/s)", "幂律阻尼浮子速度(m/s)", "浮子速度差(m/s)",
        "常阻尼振子位移(m)", "幂律阻尼振子位移(m)", "振子位移差(m)",
        "常阻尼振子速度(m/s)", "幂律阻尼振子速度(m/s)", "振子速度差(m/s)", "综合归一化差异分数",
    ], diff_detail)

    both_tail_stable = bool(stable[SCENARIOS[0]] and stable[SCENARIOS[1]])
    peak_times = [float(row[2]) for row in peak_rows]
    summary_rows = [
        ["40周期仿真尾段已形成稳定周期响应", "第35-39完整周期重复性检验",
         "关键指标最大相对偏差≤2%", "是" if both_tail_stable else "否",
         "若任一场景超过2%，则不能把末段描述为周期稳定", "误差分解/周期指标明细", "Q1结果段", ""],
        ["常阻尼与幂律阻尼均保持稳定周期响应，但稳态幅值允许存在定量差异", "阻尼结构稳健性对照",
         "两种题设结构均通过末段稳定性判据", "是" if both_stable else "否",
         "若任一阻尼结构未稳定，则该定性对比需回退重算", "结构稳健性/稳态结构差异", "Q1比较讨论", ""],
        ["局部放大必须指向两种阻尼真实差异最集中的时间区间", "全时域差异扫描 + 综合滑窗",
         f"共同ROI={roi_start:.1f}--{roi_end:.1f}s；四状态峰值时刻范围={min(peak_times):.1f}--{max(peak_times):.1f}s", "是",
         "若后续主结果变化导致ROI变化，则Figure Evidence必须同步重取ROI", "局部差异峰值/局部放大ROI/差异时序", "Q1 Figure Evidence", "不通过截轴或平滑制造差异"],
    ]
    add_sheet(book, "结论稳定性汇总", ["核心结论", "分析方法", "稳定范围", "是否保持", "失效边界", "证据工作表", "论文位置", "说明"], summary_rows)

    disposition = "support" if both_stable else "reject"
    action = "正文可保留周期稳定性结论并报告两种阻尼的稳态差异" if both_stable else "回到solve_validate复核时域与数值设置，相关结果段保持stale"
    evidence_rows = [
        ["E1", "末5完整周期重复性检验", "40周期尾段已进入稳定周期响应", disposition,
         "两场景均满足2%周期重复性阈值" if both_stable else "至少一个场景未满足2%周期重复性阈值", action, "Q1结果段"],
        ["E2", "常阻尼/幂律阻尼结构对照", "两种题设阻尼律均保持有界稳定周期响应", disposition,
         "定性稳定性保持，稳态量值差异单独报告" if both_stable else "结构对照下定性稳定性未同时保持", action, "Q1比较讨论"],
        ["E3", "全时域差异扫描 + 2.4s综合滑窗", "Local Zoom应展示两种阻尼差异最有信息量的区间", "support",
         f"差异驱动共同ROI={roi_start:.1f}--{roi_end:.1f}s；峰值集中在{min(peak_times):.1f}--{max(peak_times):.1f}s",
         "q1_plot.m必须从分析工作簿读取ROI与差异时序，不得在MATLAB中重新挑选或重新求解", "Q1 Figure Evidence"],
    ]
    add_sheet(book, "证据处置", ["Evidence ID", "method/source", "target claim", "disposition", "key finding", "required action", "paper/figure anchor"], evidence_rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)


def main() -> None:
    script_path = Path(__file__).resolve()
    project_root = script_path.parents[1]
    source = project_root / "问题一求解" / "问题一求解结果.xlsx"
    output = project_root / "问题一求解" / "问题一结果深化分析.xlsx"
    book = require_primary_workbook(source)
    period, _ = validate_primary_evidence(book)
    data = read_simulation_detail(book)
    records = cycle_metrics(data, period)
    stability_rows, references, stable = tail_stability(records)
    structure_rows, differences, both_stable = structure_comparison(references, stable)
    peak_rows, diff_detail, roi = difference_scan(data)
    context = {
        "source_hash": sha256_file(source), "records": records, "stability_rows": stability_rows,
        "stable": stable, "structure_rows": structure_rows, "differences": differences,
        "both_stable": both_stable, "peak_rows": peak_rows, "diff_detail": diff_detail, "roi": roi,
    }
    write_analysis_workbook(output, sha256_file(script_path), context)


if __name__ == "__main__":
    main()
